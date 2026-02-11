#!/usr/bin/env python3
"""
Read INSTRUCTION.xlsx and print from/to time and 15-minute slots for each entry.

15-minute division matches the calculation sheet:
  "calculation sheet for BD and non compliance of HNPCL for OCT25.xlsx"
  (sheet "New method (2)"): each slot is a row with From and To columns,
  e.g. 2:30–2:45, 2:45–3:00, 3:00–3:15, ... (consecutive 15-min intervals).
  Day is divided as 00:00, 00:15, 00:30, 00:45, 01:00, ... 23:45.
"""

import argparse
import datetime
import sys
from pathlib import Path


def format_time(val):
    """Format time or datetime for display."""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M")
    return str(val)


def format_date(val):
    """Format date or datetime for display (DD-MMM-YYYY, e.g. 05-Nov-2025)."""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%d-%b-%Y")
    if hasattr(val, "strftime"):
        return val.strftime("%d-%b-%Y")
    return str(val)


def time_to_minutes(t):
    """Convert time or datetime to minutes since midnight (0–1439). Returns None if invalid."""
    if t is None:
        return None
    if isinstance(t, datetime.datetime):
        t = t.time()
    if isinstance(t, datetime.time):
        return t.hour * 60 + t.minute
    return None


def floor_to_15(minutes):
    """Floor minutes since midnight to previous 15-min slot (0, 15, 30, 45, ...)."""
    if minutes is None:
        return None
    return (minutes // 15) * 15


def slots_15min(from_time, to_time):
    """
    Generate 15-minute slots between from_time and to_time.
    Matches calculation sheet: each slot is (From, To) with To = From + 15 min.
    Start/end are floored to previous 15-min boundary (e.g. 8:10 → 8:00).
    Returns list of (from_str, to_str) e.g. [("8:00", "8:15"), ("8:15", "8:30"), ...].
    Handles overnight (e.g. 23:00 to 00:00).
    """
    start_min = time_to_minutes(from_time)
    end_min = time_to_minutes(to_time)
    if start_min is None or end_min is None:
        return []

    start_slot = floor_to_15(start_min)
    end_slot = floor_to_15(end_min)

    if start_slot == end_slot:
        from_str = minutes_to_time_str(start_slot)
        to_str = minutes_to_time_str((start_slot + 15) % (24 * 60))
        return [(from_str, to_str)]

    # Overnight: end is next day (e.g. 23:00 → 00:00)
    if start_slot > end_slot:
        end_slot += 24 * 60

    result = []
    m = start_slot
    while m <= end_slot:
        from_m = m % (24 * 60)
        to_m = (m + 15) % (24 * 60)
        result.append((minutes_to_time_str(from_m), minutes_to_time_str(to_m)))
        m += 15
    return result


def minutes_to_time_str(minutes):
    """Convert minutes since midnight (0–1439) to HH:MM string."""
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"


def slots_to_vertical_mini_table(slots, cell_w=6, use_outer_border=False):
    """
    Build a vertical two-column mini table (From | To) for 15-min slots.
    By default uses the main table's cell as the outline (no inner box).
    With use_outer_border=True, draws full inner table borders.
    """
    if not slots:
        return ["(none)"]
    lines = []
    inner_sep = f"{'-' * (cell_w + 2)}+{'-' * (cell_w + 2)}"   # between columns only
    if use_outer_border:
        full_sep = f"+{'-' * (cell_w + 2)}+{'-' * (cell_w + 2)}+"
        lines.append(full_sep)
    header = f" {'From':<{cell_w}} | {'To':<{cell_w}} "
    lines.append(header)
    lines.append(inner_sep)
    for from_str, to_str in slots:
        row = f" {from_str:<{cell_w}} | {to_str:<{cell_w}} "
        lines.append(row)
    if use_outer_border:
        lines.append(full_sep)
    return lines

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def find_time_columns(sheet, max_header_rows=4):
    """
    Find 'from time' and 'to time' column indices (0-based).
    Checks first max_header_rows; supports multi-row headers and merged cells.
    Returns (from_idx, to_idx, data_start_row) or (None, None, 1).
    """
    time_cols = []
    for row_num in range(1, max_header_rows + 1):
        row_cells = list(sheet.iter_rows(min_row=row_num, max_row=row_num))[0]
        for idx, cell in enumerate(row_cells):
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            if "time" in val and idx not in [c for c, _ in time_cols]:
                time_cols.append((idx, val))
    # Data starts after the last header row we scanned
    data_start = max_header_rows + 1
    if len(time_cols) >= 2:
        return time_cols[0][0], time_cols[1][0], data_start
    if len(time_cols) == 1:
        return time_cols[0][0], time_cols[0][0], data_start
    return None, None, 1


def find_date_column(sheet, max_header_rows=4):
    """Find first 'date' column index (0-based) in header rows. Returns None if not found."""
    for row_num in range(1, max_header_rows + 1):
        row_cells = list(sheet.iter_rows(min_row=row_num, max_row=row_num))[0]
        for idx, cell in enumerate(row_cells):
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            if "date" in val:
                return idx
    return None


def main():
    parser = argparse.ArgumentParser(
        description="Read INSTRUCTION.xlsx and print from/to time and 15-minute slots for each entry."
    )
    parser.add_argument(
        "directory",
        help="Path to directory containing INSTRUCTION.xlsx"
    )
    args = parser.parse_args()

    input_dir = Path(args.directory)
    if not input_dir.is_dir():
        print(f"Directory does not exist: {input_dir}", file=sys.stderr)
        sys.exit(1)

    # Search for fixed filenames in the directory
    for name in ("INSTRUCTION.xlsx", "instructions.xlsx", "Instruction.xlsx"):
        candidate = input_dir / name
        if candidate.exists():
            path = candidate
            break
    else:
        print(f"No INSTRUCTION.xlsx found in directory: {input_dir}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active

    from_idx, to_idx, data_start_row = find_time_columns(sheet)
    if from_idx is None:
        print("Could not find 'from time' / 'to time' columns in first 5 rows.", file=sys.stderr)
        wb.close()
        sys.exit(1)

    date_idx = find_date_column(sheet, max_header_rows=4)

    # Build data: one entry per instruction row (row_num, date, from, to, slots)
    # slots = list of (from_str, to_str) for each 15-min interval
    rows_data = []
    slot_cell_w = 6
    w_slots_col = 0
    w_date = 10  # YYYY-MM-DD
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        row = list(row)
        date_val = row[date_idx] if date_idx is not None and date_idx < len(row) else ""
        from_time = row[from_idx] if from_idx < len(row) else ""
        to_time = row[to_idx] if to_idx < len(row) else ""
        slots = slots_15min(from_time, to_time)
        date_str = format_date(date_val)
        fr_str = format_time(from_time)
        to_str = format_time(to_time)
        mini = slots_to_vertical_mini_table(slots, slot_cell_w)
        w_slots_col = max(w_slots_col, max(len(ln) for ln in mini))
        if date_str:
            w_date = max(w_date, len(date_str))
        rows_data.append((row_num, date_str, fr_str, to_str, slots))

    print(f"Reading: {path.name}\n")
    print(f"Found {len(rows_data)} entries\n")

    # Create output workbook
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Back down and Non compliance"

    # Define styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Main title
    output_sheet.merge_cells('A1:J1')
    title_cell = output_sheet['A1']
    title_cell.value = "Back down and Non compliance of HNPCL for Oct 2025"
    title_cell.font = title_font
    title_cell.alignment = center_align

    # Row 2: Category headers
    # "Back down" spans columns A-G (7 columns)
    output_sheet.merge_cells('A2:G2')
    backdown_header = output_sheet['A2']
    backdown_header.value = "Back down"
    backdown_header.font = header_font
    backdown_header.alignment = center_align
    backdown_header.border = thin_border

    # "Non compliance" spans columns H-J (3 columns)
    output_sheet.merge_cells('H2:J2')
    noncompliance_header = output_sheet['H2']
    noncompliance_header.value = "Non compliance"
    noncompliance_header.font = header_font
    noncompliance_header.alignment = center_align
    noncompliance_header.border = thin_border

    # Row 3: Column headers
    headers_backdown = ["Date", "From", "To", "DC (MW)", "As per SLDC Scada in MW", "Diff (MW)", "Mus"]
    headers_noncompliance = ["MW as per ramp", "Diff", "MU"]
    
    # Back down headers (columns A-G)
    for col_idx, header in enumerate(headers_backdown, start=1):
        cell = output_sheet.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        # Yellow background for "As per SLDC Scada in MW"
        if header == "As per SLDC Scada in MW":
            cell.fill = yellow_fill

    # Non compliance headers (columns H-J)
    for col_idx, header in enumerate(headers_noncompliance, start=8):
        cell = output_sheet.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows starting from row 4
    for row_idx, (row_num, date_str, fr_str, to_str, slots) in enumerate(rows_data, start=4):
        # Back down columns
        output_sheet.cell(row=row_idx, column=1).value = date_str  # Date
        output_sheet.cell(row=row_idx, column=2).value = fr_str   # From
        output_sheet.cell(row=row_idx, column=3).value = to_str   # To
        # Columns 4-7 (DC (MW), As per SLDC Scada in MW, Diff (MW), Mus) left empty
        # Columns 8-10 (MW as per ramp, Diff, MU) left empty
        
        # Apply borders to all cells in the row
        for col in range(1, 11):
            output_sheet.cell(row=row_idx, column=col).border = thin_border

    # Adjust column widths
    output_sheet.column_dimensions['A'].width = 12  # Date
    output_sheet.column_dimensions['B'].width = 8   # From
    output_sheet.column_dimensions['C'].width = 8   # To
    output_sheet.column_dimensions['D'].width = 10  # DC (MW)
    output_sheet.column_dimensions['E'].width = 25  # As per SLDC Scada in MW
    output_sheet.column_dimensions['F'].width = 12  # Diff (MW)
    output_sheet.column_dimensions['G'].width = 10  # Mus
    output_sheet.column_dimensions['H'].width = 15  # MW as per ramp
    output_sheet.column_dimensions['I'].width = 10  # Diff
    output_sheet.column_dimensions['J'].width = 10  # MU

    # Set row heights
    output_sheet.row_dimensions[1].height = 25  # Title row
    output_sheet.row_dimensions[2].height = 20  # Category headers
    output_sheet.row_dimensions[3].height = 30  # Column headers (for wrapped text)

    # Save output file
    output_path = input_dir / "Back_down_and_Non_compliance_output.xlsx"
    output_wb.save(output_path)
    print(f"Output saved to: {output_path}")
    print(f"Created {len(rows_data)} data rows")

    wb.close()


if __name__ == "__main__":
    main()
