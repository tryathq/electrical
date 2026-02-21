"""Build report Excel workbook from output rows."""

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# Visible columns + hidden marker column
HEADERS = ["Date", "From", "To", "Ramp rate", "DC (MW)", "As per SLDC Scada in MW", "DC , Scada Diff (MW)", "Mus", "Sum Mus", "MW as per ramp", "Diff", "MU", "Sum MU", "_ins_end"]
COLUMN_WIDTHS = [15, 10, 10, 12, 12, 25, 12, 12, 12, 14, 12, 12, 12, 8]
PAD = 0

# Yellow highlight fill
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _make_styles():
    return {
        "header_font": Font(bold=True, size=11),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "thin_border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def build_report_workbook(output_rows: list[dict]) -> Workbook:
    """
    Build and return an openpyxl Workbook with 'Time Intervals' sheet
    filled with output_rows. Caller should save to path.
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Time Intervals"
    styles = _make_styles()
    header_font = styles["header_font"]
    center_align = styles["center_align"]
    thin_border = styles["thin_border"]

    header_row = 1 + PAD
    start_col = 1 + PAD
    start_data_row = 2 + PAD

    # Header row
    for c, h in enumerate(HEADERS):
        cell = sheet.cell(row=header_row, column=start_col + c)
        cell.value = h
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    def _is_summary_row(row_dict: dict) -> bool:
        """Row that only has Sum Mus filled (instruction-entry sum)."""
        return (
            not row_dict.get("Date") and not row_dict.get("From") and not row_dict.get("To")
            and row_dict.get("Sum Mus") != "" and row_dict.get("Sum Mus") is not None
        )

    # Data rows and date merging
    date_start_row: Optional[int] = None
    row_idx = start_data_row
    for r, row_dict in enumerate(output_rows):
        out_row = row_idx + r
        if row_dict.get("Date") and date_start_row is not None and out_row > date_start_row:
            merge_end = out_row - 1
            # Exclude summary row from merge so that row keeps empty Date/From/To
            if r > 0 and _is_summary_row(output_rows[r - 1]):
                merge_end = out_row - 2
            if merge_end >= date_start_row:
                sheet.merge_cells(
                    f"{get_column_letter(start_col)}{date_start_row}:{get_column_letter(start_col)}{merge_end}"
                )
        if row_dict.get("Date"):
            date_start_row = out_row
        # Get values
        date_val = row_dict.get("Date") or ""
        to_val = row_dict.get("To")
        ins_end = row_dict.get("_ins_end", False)
        
        # Write cell values
        date_cell = sheet.cell(row=out_row, column=start_col)
        date_cell.value = date_val
        
        sheet.cell(row=out_row, column=start_col + 1).value = row_dict.get("From")
        
        to_cell = sheet.cell(row=out_row, column=start_col + 2)
        to_cell.value = to_val
        
        ramp_rate_cell = sheet.cell(row=out_row, column=start_col + 3)
        ramp_rate_cell.value = row_dict.get("Ramp rate", "")
        ramp_rate_note = row_dict.get("Ramp rate_note", "")
        if not ramp_rate_note and isinstance(row_dict.get("tooltips"), dict):
            ramp_rate_note = row_dict["tooltips"].get("Ramp rate", "")
        if isinstance(ramp_rate_note, str) and ramp_rate_note.strip():
            ramp_rate_cell.comment = Comment(ramp_rate_note.strip(), "App")
        
        sheet.cell(row=out_row, column=start_col + 4).value = row_dict.get("DC (MW)")
        sheet.cell(row=out_row, column=start_col + 5).value = row_dict.get("As per SLDC Scada in MW")
        sheet.cell(row=out_row, column=start_col + 6).value = row_dict.get("DC , Scada Diff (MW)")
        sheet.cell(row=out_row, column=start_col + 7).value = row_dict.get("Mus")
        sheet.cell(row=out_row, column=start_col + 8).value = row_dict.get("Sum Mus")
        sheet.cell(row=out_row, column=start_col + 9).value = row_dict.get("MW as per ramp")
        sheet.cell(row=out_row, column=start_col + 10).value = row_dict.get("Diff")
        sheet.cell(row=out_row, column=start_col + 11).value = row_dict.get("MU")
        sheet.cell(row=out_row, column=start_col + 12).value = row_dict.get("Sum MU")
        
        # Write _ins_end marker (hidden column)
        ins_end_cell = sheet.cell(row=out_row, column=start_col + 13)
        ins_end_cell.value = "TRUE" if ins_end else "FALSE"
        
        # Apply yellow highlighting
        # Date column: highlight when it has a value (first row of each date/instruction)
        if date_val and str(date_val).strip():
            date_cell.fill = YELLOW_FILL
            date_cell.font = Font(bold=True)
        
        # To column: highlight when it's instruction end time
        if ins_end:
            to_cell.fill = YELLOW_FILL
            to_cell.font = Font(bold=True)
        
        # Apply borders to all columns including _ins_end
        for c in range(14):
            sheet.cell(row=out_row, column=start_col + c).border = thin_border

    if date_start_row is not None:
        last_data_row = row_idx + len(output_rows) - 1
        merge_end = last_data_row
        if output_rows and _is_summary_row(output_rows[-1]):
            merge_end = last_data_row - 1  # Exclude final summary row from merge
        if merge_end > date_start_row:
            sheet.merge_cells(
                f"{get_column_letter(start_col)}{date_start_row}:{get_column_letter(start_col)}{merge_end}"
            )

    last_row = row_idx + len(output_rows) - 1
    last_content_col = start_col + 13  # Including _ins_end column

    sheet.freeze_panes = sheet.cell(row=start_data_row, column=start_col).coordinate
    sheet.sheet_view.showGridLines = False

    for i, w in enumerate(COLUMN_WIDTHS):
        sheet.column_dimensions[get_column_letter(start_col + i)].width = w
    
    # Hide the _ins_end marker column
    sheet.column_dimensions[get_column_letter(start_col + 13)].hidden = True

    # Print area excludes the hidden _ins_end column
    sheet.print_area = f"A1:{get_column_letter(start_col + 12)}{last_row}"

    return wb
