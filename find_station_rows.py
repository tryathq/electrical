#!/usr/bin/env python3
"""
Find rows in XLSX file where "Name of the station" column matches given station name.

Usage:
  python find_station_rows.py <xlsx_path> <station_name> [options]

Example:
  python find_station_rows.py "input/Back_Down_Instructions.xlsx" HINDUJA
  python find_station_rows.py "input/jan 2026.xlsx" HINDUJA --sheet HNPCL
  python find_station_rows.py "input/instructions.xlsx" HINDUJA --dc-file "input/dc_data.xlsx"
"""

import argparse
import sys
import re
from pathlib import Path
from datetime import datetime, date, time

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_value(val):
    """Format cell value for display, especially dates and times."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        # Check if it's actually a time (1900-01-01 date with time)
        if val.date() == date(1900, 1, 1):
            return val.strftime("%H:%M")
        return val.strftime("%d-%b-%Y")
    if isinstance(val, date):
        return val.strftime("%d-%b-%Y")
    # Check if it's a time object
    if isinstance(val, time):
        return val.strftime("%H:%M")
    # Check if string looks like a time (HH:MM:SS or HH:MM)
    val_str = str(val)
    if ":" in val_str and len(val_str) <= 8:
        # Try to parse and format as time
        try:
            if len(val_str.split(":")) == 3:
                # HH:MM:SS format
                parts = val_str.split(":")
                return f"{parts[0]}:{parts[1]}"
            return val_str
        except:
            pass
    return str(val)


def time_to_minutes(t):
    """Convert time or datetime to minutes since midnight (0–1439). Returns None if invalid."""
    if t is None:
        return None
    if isinstance(t, datetime):
        t = t.time()
    if isinstance(t, time):
        return t.hour * 60 + t.minute
    # Try parsing string
    if isinstance(t, str):
        try:
            parts = t.strip().split(":")
            if len(parts) >= 2:
                h = int(parts[0])
                m = int(parts[1])
                if 0 <= h < 24 and 0 <= m < 60:
                    return h * 60 + m
        except (ValueError, IndexError):
            pass
    return None


def floor_to_15(minutes):
    """Floor minutes since midnight to previous 15-min slot (0, 15, 30, 45, ...)."""
    if minutes is None:
        return None
    return (minutes // 15) * 15


def minutes_to_time_str(minutes):
    """Convert minutes since midnight (0–1439) to HH:MM string."""
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"


def slots_15min(from_time, to_time):
    """
    Generate 15-minute slots between from_time and to_time.
    Start is floored to previous 15-min boundary (e.g. 8:10 → 8:00).
    End time is respected - slots do not extend beyond the given end time.
    Returns list of (from_str, to_str) e.g. [("8:00", "8:15"), ("8:15", "8:30"), ...].
    Handles overnight (e.g. 23:00 to 00:00).
    """
    start_min = time_to_minutes(from_time)
    end_min = time_to_minutes(to_time)
    if start_min is None or end_min is None:
        return []

    start_slot = floor_to_15(start_min)
    
    # Overnight: end is next day (e.g. 23:00 → 00:00)
    if start_slot > end_min:
        end_min += 24 * 60

    result = []
    m = start_slot
    # Only include slots where the slot's START time is < end_min
    # This ensures slots don't extend beyond the end time
    while m < end_min:
        from_m = m % (24 * 60)
        to_m = (m + 15) % (24 * 60)
        result.append((minutes_to_time_str(from_m), minutes_to_time_str(to_m)))
        m += 15
    return result


def parse_time_str(time_str):
    """Parse HH:MM string to minutes since midnight. Returns None if invalid."""
    return time_to_minutes(time_str)


def convert_date_to_sheet_format(date_str):
    """
    Convert date string from format '02-Jan-2026' to sheet format '02.01.2026'.
    Handles various date formats and returns None if parsing fails.
    """
    if not date_str:
        return None
    
    date_str = str(date_str).strip()
    
    # List of formats to try
    formats = [
        "%d-%b-%Y",      # 02-Jan-2026
        "%d-%b-%y",      # 02-Jan-26
        "%d.%m.%Y",      # 02.01.2026 (already in correct format)
        "%d/%m/%Y",      # 02/01/2026
        "%Y-%m-%d",      # 2026-01-02
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            # Convert to DD.MM.YYYY format
            result = dt.strftime("%d.%m.%Y")
            return result
        except ValueError:
            continue
    
    return None


def normalize_time_str(time_str):
    """
    Normalize time string to HH:MM format for comparison.
    Handles formats like "0:00", "00:00", "1:15", "01:15", etc.
    """
    if not time_str:
        return ""
    time_str = str(time_str).strip()
    # Parse and reformat to ensure consistent HH:MM format
    minutes = time_to_minutes(time_str)
    if minutes is not None:
        return minutes_to_time_str(minutes)
    return time_str


def convert_date_for_bd_filename(date_str):
    """
    Convert date string from format '01-Jan-2026' to various formats for BD filename matching.
    Returns list of possible date strings to search for in filenames.
    """
    if not date_str:
        return []
    
    date_str = str(date_str).strip()
    date_formats = [
        "%d-%b-%Y",      # 01-Jan-2026
        "%d-%b-%y",      # 01-Jan-26
        "%d.%m.%Y",      # 01.01.2026
        "%d/%m/%Y",      # 01/01/2026
        "%Y-%m-%d",      # 2026-01-01
    ]
    
    possible_dates = []
    for fmt in date_formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            # Generate various formats for filename matching
            day_no_zero = str(dt.day)  # Day without leading zero
            month_no_zero = str(dt.month)  # Month without leading zero
            possible_dates.extend([
                dt.strftime("%d/%m/%Y"),      # 01/01/2026
                f"{day_no_zero}/{month_no_zero}/{dt.year}",  # 1/1/2026 (no leading zeros)
                dt.strftime("%d-%m-%Y"),      # 01-01-2026
                f"{day_no_zero}-{month_no_zero}-{dt.year}",  # 1-1-2026
                dt.strftime("%Y-%m-%d"),      # 2026-01-01
                dt.strftime("%Y/%m/%d"),      # 2026/01/01
            ])
        except ValueError:
            continue
    
    # Remove duplicates while preserving order
    seen = set()
    unique_dates = []
    for date in possible_dates:
        if date not in seen:
            seen.add(date)
            unique_dates.append(date)
    
    return unique_dates


def find_bd_file(bd_folder, date_str):
    """
    Find BD file in folder that contains the given date in its filename.
    Returns Path to file or None if not found.
    """
    if not bd_folder or not bd_folder.exists() or not bd_folder.is_dir():
        return None
    
    if not date_str:
        return None
    
    # Get possible date formats for filename matching
    possible_dates = convert_date_for_bd_filename(date_str)
    
    # Search for files containing any of the date formats
    for file_path in bd_folder.glob("*.xlsx"):
        filename_lower = file_path.name.lower()
        for date_format in possible_dates:
            if date_format.lower() in filename_lower:
                return file_path
    
    # Also try .xls files
    for file_path in bd_folder.glob("*.xls"):
        filename_lower = file_path.name.lower()
        for date_format in possible_dates:
            if date_format.lower() in filename_lower:
                return file_path
    
    return None


class SCADALookupCache:
    """Cache for BD file lookups - maintains file list and loads files on demand."""
    def __init__(self, bd_folder, column_name, sheet_name=None):
        self.bd_folder = bd_folder
        self.column_name = column_name
        self.sheet_name = sheet_name  # Specific sheet to read (e.g., "DATA-CMD")
        self.cache = {}  # {date_str: (wb, ws, time_col, target_col, header_row, time_map)}
        self.file_list = []  # List of (file_path, possible_dates) tuples
        self.column_cache = {}  # {file_path: (time_col, target_col, header_row)}
        
        # Build file list at initialization (just paths, no opening)
        self._build_file_list()
    
    def _build_file_list(self):
        """Build list of BD files with their possible date matches (no file opening)."""
        if not self.bd_folder or not self.bd_folder.exists():
            return
        
        # Get all Excel files in BD folder
        excel_files = list(self.bd_folder.glob("*.xlsx")) + list(self.bd_folder.glob("*.xls"))
        
        for file_path in excel_files:
            filename_lower = file_path.name.lower()
            # Extract possible dates from filename
            possible_dates = []
            
            # Try to extract dates from filename (common patterns)
            # Pattern for dates like 1/1/2026, 01/01/2026, 2026-01-01, etc.
            date_patterns = [
                r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',  # 1/1/2026 or 01-01-2026
                r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',  # 2026-01-01
            ]
            
            for pattern in date_patterns:
                matches = re.findall(pattern, filename_lower)
                for match in matches:
                    if len(match) == 3:
                        # Try to construct date string
                        if len(match[2]) == 4:  # YYYY format
                            month, day, year = match
                            possible_dates.append(f"{day}/{month}/{year}")
                            possible_dates.append(f"{int(day)}/{int(month)}/{year}")
                        else:  # DD/MM/YYYY format
                            day, month, year = match
                            possible_dates.append(f"{day}/{month}/{year}")
                            possible_dates.append(f"{int(day)}/{int(month)}/{year}")
            
            if possible_dates:
                self.file_list.append((file_path, possible_dates))
    
    def _find_file_for_date(self, date_str):
        """Find BD file for given date from pre-built file list."""
        # Convert date to possible formats for matching
        possible_date_formats = convert_date_for_bd_filename(date_str)
        
        # Check each file's possible dates
        for file_path, file_dates in self.file_list:
            filename_lower = file_path.name.lower()
            # Check if any of our date formats match file's dates or filename
            for date_format in possible_date_formats:
                if date_format.lower() in filename_lower:
                    return file_path
            # Also check file's extracted dates
            for file_date in file_dates:
                for date_format in possible_date_formats:
                    if date_format.lower() == file_date.lower():
                        return file_path
        
        return None
    
    def get_workbook_for_date(self, date_str, show_progress=False):
        """Get or load workbook for given date (loads file only when needed)."""
        if date_str in self.cache:
            return self.cache[date_str]
        
        # Find BD file from pre-built list
        bd_file = self._find_file_for_date(date_str)
        if not bd_file:
            if show_progress:
                print(f" (file not found)", end="", flush=True)
            return None
        
        if show_progress:
            print(f" (loading {bd_file.name})", end="", flush=True)
        
        
        try:
            # Check if we already have column info for this file
            if bd_file in self.column_cache:
                time_col, target_col, header_row = self.column_cache[bd_file]
                # Load workbook (we already know columns, so just load once)
                wb = openpyxl.load_workbook(bd_file, read_only=True, data_only=True)
                
                # Get the specified sheet or active sheet
                if self.sheet_name:
                    sheet_found = None
                    sheet_name_lower = self.sheet_name.lower().strip()
                    for name in wb.sheetnames:
                        if name.strip().lower() == sheet_name_lower or sheet_name_lower in name.strip().lower():
                            sheet_found = name
                            break
                    if sheet_found:
                        ws = wb[sheet_found]
                    else:
                        return None
                else:
                    ws = wb.active
            else:
                # Load workbook ONCE to find columns AND use it
                wb = openpyxl.load_workbook(bd_file, read_only=True, data_only=True)
                
                # Get the specified sheet or active sheet
                if self.sheet_name:
                    # Try to find sheet by name (case-insensitive, flexible matching)
                    sheet_found = None
                    sheet_name_lower = self.sheet_name.lower().strip()
                    for name in wb.sheetnames:
                        if name.strip().lower() == sheet_name_lower or sheet_name_lower in name.strip().lower():
                            sheet_found = name
                            break
                    if sheet_found:
                        ws = wb[sheet_found]
                    else:
                        wb.close()
                        return None
                else:
                    ws = wb.active
                
                # Find columns
                time_col = None
                target_col = None
                header_row = None
                
                for row_num in range(1, min(11, ws.max_row + 1)):
                    for col_idx in range(1, min(ws.max_column + 1, 50)):
                        cell = ws.cell(row=row_num, column=col_idx)
                        if cell.value:
                            header_val = str(cell.value).strip().lower()
                            if "time" in header_val and time_col is None:
                                time_col = col_idx
                                if header_row is None:
                                    header_row = row_num
                            if self.column_name.lower() in header_val or header_val in self.column_name.lower():
                                target_col = col_idx
                                if header_row is None:
                                    header_row = row_num
                    
                    if time_col and target_col:
                        break
                
                if not time_col or not target_col:
                    wb.close()
                    return None
                
                # Cache column info for this file (for future dates using same file)
                self.column_cache[bd_file] = (time_col, target_col, header_row)
            
            # Build time-to-row mapping efficiently (limit to reasonable number)
            time_map = {}
            data_start = header_row + 1
            # Limit to 96 rows per day (15-min intervals * 24 hours) + buffer
            max_rows = min(ws.max_row + 1, data_start + 150)
            
            for row_num in range(data_start, max_rows):
                time_cell = ws.cell(row=row_num, column=time_col)
                if time_cell.value is None:
                    continue
                
                time_cell_raw = time_cell.value
                
                # Handle different time formats
                time_norm = None
                
                # If it's a datetime object, extract time directly
                if isinstance(time_cell_raw, datetime):
                    time_norm = normalize_time_str(time_cell_raw.strftime("%H:%M:%S"))
                else:
                    time_cell_val = format_value(time_cell_raw)
                    
                    # Extract time from various formats:
                    # "01/01/2026 20:15:00" -> "20:15"
                    # "20:15:00" -> "20:15"
                    # "20:15" -> "20:15"
                    
                    # If it contains a space, likely has date and time
                    if " " in str(time_cell_val):
                        parts = str(time_cell_val).split()
                        # Get the last part which should be time
                        if len(parts) > 1:
                            time_part = parts[-1]
                            time_norm = normalize_time_str(time_part)
                    else:
                        # Just time, normalize directly
                        time_norm = normalize_time_str(time_cell_val)
                
                if time_norm:
                    time_map[time_norm] = row_num
                    # Also store variations (with/without seconds)
                    if ":" in time_norm:
                        time_parts = time_norm.split(":")
                        if len(time_parts) == 3:  # Has seconds
                            time_no_sec = f"{time_parts[0]}:{time_parts[1]}"
                            if time_no_sec not in time_map:
                                time_map[time_no_sec] = row_num
            
            cache_entry = (wb, ws, time_col, target_col, header_row, time_map)
            self.cache[date_str] = cache_entry
            
            return cache_entry
            
        except Exception as e:
            if show_progress:
                print(f" (error)", end="", flush=True)
            return None
    
    def find_value(self, date_str, time_str, debug=False, show_progress=False):
        """Find SCADA value for given date and time."""
        cache_entry = self.get_workbook_for_date(date_str, show_progress=show_progress)
        if not cache_entry:
            return None
        
        wb, ws, time_col, target_col, header_row, time_map = cache_entry
        
        # Normalize time for lookup
        time_norm = normalize_time_str(time_str)
        
        # Look up row number from cache
        row_num = time_map.get(time_norm)
        if row_num:
            target_cell = ws.cell(row=row_num, column=target_col)
            return target_cell.value
        
        # Fallback: search if not in cache (limit search range)
        data_start = header_row + 1
        max_search = min(ws.max_row + 1, data_start + 200)
        for row_num in range(data_start, max_search):
            time_cell = ws.cell(row=row_num, column=time_col)
            if time_cell.value is None:
                continue
            
            time_cell_raw = time_cell.value
            
            # Extract time part if it's a datetime string
            time_cell_norm = None
            if isinstance(time_cell_raw, datetime):
                time_cell_norm = normalize_time_str(time_cell_raw.strftime("%H:%M:%S"))
            else:
                time_cell_val = format_value(time_cell_raw)
                if " " in str(time_cell_val):
                    parts = str(time_cell_val).split()
                    if len(parts) > 1:
                        time_part = parts[-1]
                        time_cell_norm = normalize_time_str(time_part)
                else:
                    time_cell_norm = normalize_time_str(time_cell_val)
            
            # Match normalized times
            if time_cell_norm == time_norm or time_norm in str(time_cell_val) or time_norm in str(time_cell_raw):
                target_cell = ws.cell(row=row_num, column=target_col)
                return target_cell.value
        
        return None
    
    def close_all(self):
        """Close all cached workbooks."""
        for cache_entry in self.cache.values():
            if cache_entry:
                wb = cache_entry[0]
                try:
                    wb.close()
                except:
                    pass
        self.cache.clear()
        self.column_cache.clear()


def find_scada_value(scada_cache, date_str, time_str, debug=False, show_progress=False):
    """
    Find SCADA value using cached lookup.
    """
    if not scada_cache:
        return None
    
    return scada_cache.find_value(date_str, time_str, debug=debug, show_progress=show_progress)


def find_dc_value(dc_wb, sheet_name, from_time_str, to_time_str, debug=False):
    """
    Find DC value from DC workbook sheet for matching time range.
    Searches for row where 'From' and 'To' columns match the given time range.
    Returns the 'Final Revison' column value, or None if not found.
    """
    if dc_wb is None:
        if debug:
            print(f"  [DC Lookup] dc_wb is None", file=sys.stderr)
        return None
    
    if not sheet_name:
        if debug:
            print(f"  [DC Lookup] sheet_name is empty", file=sys.stderr)
        return None
    
    # Find the sheet by name (case-insensitive, flexible matching)
    target_sheet = None
    sheet_name_lower = sheet_name.lower().strip()
    
    # Try exact match first (with date format)
    for name in dc_wb.sheetnames:
        name_clean = name.strip()
        if name_clean.lower() == sheet_name_lower:
            target_sheet = name
            break
    
    # If not found, try partial match (date might be embedded in sheet name)
    if target_sheet is None:
        for name in dc_wb.sheetnames:
            name_clean = name.strip().lower()
            # Check if the date string is contained in the sheet name
            if sheet_name_lower in name_clean:
                target_sheet = name
                break
    
    if target_sheet is None:
        if debug:
            print(f"  [DC Lookup] Sheet '{sheet_name}' not found in DC file.", file=sys.stderr)
            print(f"  [DC Lookup] Available sheets ({len(dc_wb.sheetnames)}): {', '.join(dc_wb.sheetnames[:5])}...", file=sys.stderr)
        return None
    
    if debug:
        print(f"  [DC Lookup] Found sheet: '{target_sheet}'", file=sys.stderr)
    
    ws = dc_wb[target_sheet]
    
    # Find column indices for "From", "To", and "Final Revison"
    from_col = None
    to_col = None
    final_revision_col = None
    header_row = None
    
    # First pass: find the header row by looking for a row that has multiple expected headers
    for row_num in range(1, min(11, ws.max_row + 1)):
        found_headers = []
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value:
                header_val = str(cell.value).strip().lower()
                if "from" in header_val:
                    found_headers.append(("from", col_idx))
                elif "to" in header_val and "tb" not in header_val and "no" not in header_val:
                    found_headers.append(("to", col_idx))
                elif "final" in header_val and ("revis" in header_val or "revisi" in header_val):
                    found_headers.append(("final", col_idx))
        
        # If we found at least 2 of the expected headers in this row, use it as header row
        if len(found_headers) >= 2:
            header_row = row_num
            for header_type, col_idx in found_headers:
                if header_type == "from" and from_col is None:
                    from_col = col_idx
                elif header_type == "to" and to_col is None:
                    to_col = col_idx
                elif header_type == "final" and final_revision_col is None:
                    final_revision_col = col_idx
            break
    
    # If we didn't find a good header row, do a second pass with less strict matching
    if header_row is None:
        for row_num in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value:
                    header_val = str(cell.value).strip().lower()
                    if "from" in header_val and from_col is None:
                        from_col = col_idx
                        header_row = row_num
                    elif "to" in header_val and to_col is None and col_idx != from_col:
                        if "tb" not in header_val and "no" not in header_val:
                            to_col = col_idx
                            if header_row is None:
                                header_row = row_num
                    elif ("final" in header_val and ("revis" in header_val or "revisi" in header_val)) and final_revision_col is None:
                        final_revision_col = col_idx
                        if header_row is None:
                            header_row = row_num
    
    if not (from_col and to_col and final_revision_col):
        if debug:
            print(f"  [DC Lookup] Required columns not found (From={from_col}, To={to_col}, Final={final_revision_col})", file=sys.stderr)
        return None
    
    if debug:
        print(f"  [DC Lookup] Columns found: From={from_col}, To={to_col}, Final={final_revision_col}", file=sys.stderr)
    
    # Normalize time strings for comparison
    from_time_norm = normalize_time_str(from_time_str)
    to_time_norm = normalize_time_str(to_time_str)
    
    # Search for matching time range
    data_start = header_row + 1
    max_rows_to_check = min(ws.max_row + 1, data_start + 200)
    
    if debug:
        print(f"  [DC Lookup] Searching for time range: {from_time_norm} - {to_time_norm}", file=sys.stderr)
    
    for row_num in range(data_start, max_rows_to_check):
        from_cell = ws.cell(row=row_num, column=from_col)
        to_cell = ws.cell(row=row_num, column=to_col)
        
        if from_cell.value is None or to_cell.value is None:
            continue
        
        # Format and normalize cell values for comparison
        from_val = normalize_time_str(format_value(from_cell.value))
        to_val = normalize_time_str(format_value(to_cell.value))
        
        # Match time range
        if from_val == from_time_norm and to_val == to_time_norm:
            dc_cell = ws.cell(row=row_num, column=final_revision_col)
            if debug:
                print(f"  [DC Lookup] Match found at row {row_num}: DC value = {dc_cell.value}", file=sys.stderr)
            return dc_cell.value
    
    if debug:
        print(f"  [DC Lookup] No match found for {from_time_norm} - {to_time_norm}", file=sys.stderr)
    
    return None


def find_column_by_name(ws, column_name, max_header_rows=10):
    """
    Search for a column in the sheet by header name (case-insensitive, partial match).
    Scans first max_header_rows for header row.
    Returns (1-based column index, header_row) or (None, None).
    """
    target = column_name.strip().lower()
    if not target:
        return None, None
    
    for row_num in range(1, min(max_header_rows + 1, ws.max_row + 1)):
        for col_idx, cell in enumerate(ws[row_num], start=1):
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            # Exact match first
            if val == target:
                return col_idx, row_num
            # Partial match (target contained in header or header contained in target)
            if target in val or val in target:
                return col_idx, row_num
    
    return None, None


def find_matching_rows(ws, station_col_idx, station_name, header_row):
    """
    Find all rows where the station column matches the given station name.
    Returns list of (row_num, row_data) tuples.
    """
    target = station_name.strip()
    matches = []
    data_start = (header_row or 1) + 1
    
    for row_num in range(data_start, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=station_col_idx)
        if cell.value is None:
            continue
        
        cell_val = str(cell.value).strip()
        # Exact match
        if cell_val == target:
            row_data = [ws.cell(row=row_num, column=c).value 
                       for c in range(1, ws.max_column + 1)]
            matches.append((row_num, row_data))
        # Case-insensitive match
        elif cell_val.upper() == target.upper():
            row_data = [ws.cell(row=row_num, column=c).value 
                       for c in range(1, ws.max_column + 1)]
            matches.append((row_num, row_data))
        # Partial match (station name contained in cell value)
        elif target.upper() in cell_val.upper():
            row_data = [ws.cell(row=row_num, column=c).value 
                       for c in range(1, ws.max_column + 1)]
            matches.append((row_num, row_data))
    
    return matches


def main():
    parser = argparse.ArgumentParser(
        description="Find rows in XLSX file where 'Name of the station' column matches given station name."
    )
    parser.add_argument(
        "xlsx_path",
        type=Path,
        help="Path to the XLSX file",
    )
    parser.add_argument(
        "station_name",
        help="Station name to search for (e.g., HINDUJA)",
    )
    parser.add_argument(
        "--sheet",
        help="Sheet name to search in (default: first/active sheet)",
        default=None,
    )
    parser.add_argument(
        "--column",
        help="Column name to search (default: searches for 'Name of the station')",
        default="Name of the station",
    )
    parser.add_argument(
        "--header-rows",
        type=int,
        default=10,
        help="Max header rows to scan for column name (default: 10)",
    )
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="Read with data_only=True (evaluated values, not formulas)",
    )
    parser.add_argument(
        "--max-columns",
        type=int,
        default=20,
        help="Max columns to display per row (default: 20)",
    )
    parser.add_argument(
        "--dc-file",
        type=Path,
        help="Path to DC Excel file with date-named sheets (e.g., '01.01.2026')",
        default=None,
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose debug output for DC lookup",
    )
    parser.add_argument(
        "--bd-folder",
        type=Path,
        help="Path to BD folder containing SCADA files (default: data/BD)",
        default=None,
    )
    parser.add_argument(
        "--scada-column",
        help="Column header name to extract from BD files (e.g., 'HNJA4_AG.STTN.X_BUS_GEN.MW')",
        default=None,
    )
    parser.add_argument(
        "--bd-sheet",
        help="Sheet name to read from BD files (e.g., 'DATA-CMD'). If not specified, uses active sheet.",
        default=None,
    )
    
    args = parser.parse_args()
    
    xlsx_path = args.xlsx_path
    if not xlsx_path.is_file():
        print(f"Error: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    
    # Resolve BD folder path
    bd_folder = None
    if args.bd_folder:
        bd_folder = args.bd_folder
        if not bd_folder.is_absolute():
            # Try relative to input file directory first
            bd_folder_relative = xlsx_path.parent / bd_folder
            if bd_folder_relative.exists() and bd_folder_relative.is_dir():
                bd_folder = bd_folder_relative
            elif not bd_folder.exists():
                # Try current directory
                bd_folder_cwd = Path.cwd() / bd_folder
                if bd_folder_cwd.exists() and bd_folder_cwd.is_dir():
                    bd_folder = bd_folder_cwd
                else:
                    print(f"Warning: BD folder not found: {args.bd_folder}", file=sys.stderr)
                    bd_folder = None
        
        if bd_folder and bd_folder.exists() and bd_folder.is_dir():
            pass  # BD folder found, continue
        elif bd_folder:
            bd_folder = None
    elif args.scada_column:
        # If scada-column is provided but no bd-folder, try default
        default_bd = Path("data/BD")
        if default_bd.exists() and default_bd.is_dir():
            bd_folder = default_bd
        else:
            bd_folder = None
    
    if args.scada_column and not bd_folder:
        print("Warning: BD folder not found. SCADA values will not be filled.", file=sys.stderr)
    
    # Load DC workbook if provided
    dc_wb = None
    if args.dc_file:
        # Try to resolve the path - check multiple locations
        dc_path = args.dc_file
        dc_path_resolved = None
        
        # List of paths to try
        paths_to_try = []
        
        if dc_path.is_absolute():
            # Absolute path - try as-is
            paths_to_try.append(dc_path)
        else:
            # Relative paths - try multiple locations
            # 1. Relative to input file directory
            paths_to_try.append(xlsx_path.parent / dc_path)
            # 2. Relative to current working directory
            paths_to_try.append(Path.cwd() / dc_path)
            # 3. Just the filename in input file directory
            if dc_path.name != str(dc_path):
                paths_to_try.append(xlsx_path.parent / dc_path.name)
            # 4. Just the filename in current directory
            if dc_path.name != str(dc_path):
                paths_to_try.append(Path.cwd() / dc_path.name)
        
        # Try each path
        for path_attempt in paths_to_try:
            if path_attempt.is_file():
                dc_path_resolved = path_attempt
                break
        
        if dc_path_resolved and dc_path_resolved.is_file():
            try:
                dc_wb = openpyxl.load_workbook(dc_path_resolved, read_only=True, data_only=True)
            except Exception as e:
                print(f"Error loading DC file: {e}", file=sys.stderr)
                print("Continuing without DC values...", file=sys.stderr)
    
    # Load workbook
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=args.data_only)
    
    # Select sheet
    if args.sheet:
        # Search for sheet by name
        sheet_found = None
        target = args.sheet.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower() == target or target in name.strip().lower():
                sheet_found = name
                break
        if sheet_found is None:
            print(f"Error: No sheet matching '{args.sheet}' in {xlsx_path.name}", file=sys.stderr)
            print(f"Available sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
            wb.close()
            sys.exit(1)
        ws = wb[sheet_found]
    else:
        ws = wb.active
    
    # Find column
    col_idx, header_row = find_column_by_name(ws, args.column, max_header_rows=args.header_rows)
    if col_idx is None:
        print(f"Error: No column matching '{args.column}' found in sheet '{ws.title}'", file=sys.stderr)
        print(f"Searched first {args.header_rows} rows.", file=sys.stderr)
        print(f"\nFirst row sample:", file=sys.stderr)
        first_row = [str(cell.value)[:30] if cell.value else "" for cell in ws[1][:15]]
        print(f"  {first_row}", file=sys.stderr)
        wb.close()
        sys.exit(1)
    
    # Find matching rows
    matches = find_matching_rows(ws, col_idx, args.station_name, header_row)
    
    # Try to find "From Time" and "To Time" columns for 15-minute extraction
    from_time_col = None
    to_time_col = None
    date_col = None
    
    for col_idx_header in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=header_row, column=col_idx_header)
        if header_cell.value:
            header_val = str(header_cell.value).strip().lower()
            if "from" in header_val and "time" in header_val:
                from_time_col = col_idx_header
            elif "to" in header_val and "time" in header_val:
                to_time_col = col_idx_header
            elif "date" in header_val:
                date_col = col_idx_header
    
    if not date_col and (args.dc_file or args.scada_column):
        print("Warning: Date column not found.", file=sys.stderr)
    if not from_time_col or not to_time_col:
        print("Warning: From/To Time columns not found.", file=sys.stderr)
    
    if not matches:
        print(f"No rows found where '{args.column}' = '{args.station_name}'")
        wb.close()
        sys.exit(0)
    
    # Create output Excel file
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Time Intervals"
    
    # Define styles
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    output_sheet['A1'] = 'Date'
    output_sheet['B1'] = 'From'
    output_sheet['C1'] = 'To'
    output_sheet['D1'] = 'DC (MW)'
    output_sheet['E1'] = 'As per SLDC Scada in MW'
    
    # Apply header styles
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        cell = output_sheet[col]
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Initialize SCADA cache if BD folder is provided (builds file list, loads files on demand)
    scada_cache = None
    if bd_folder and args.scada_column:
        scada_cache = SCADALookupCache(bd_folder, args.scada_column, args.bd_sheet)
    
    # Populate data rows
    row_idx = 2
    dc_found_count = 0
    dc_not_found_count = 0
    scada_found_count = 0
    scada_not_found_count = 0
    
    # Track progress for SCADA lookups
    total_slots = 0
    processed_slots = 0
    current_date = None
    
    # Count total slots first (for progress calculation)
    for idx, (row_num, row_data) in enumerate(matches, 1):
        if from_time_col and to_time_col and from_time_col <= len(row_data) and to_time_col <= len(row_data):
            from_time_val = row_data[from_time_col - 1] if from_time_col > 0 else None
            to_time_val = row_data[to_time_col - 1] if to_time_col > 0 else None
            if from_time_val is not None and to_time_val is not None:
                slots = slots_15min(from_time_val, to_time_val)
                total_slots += len(slots) if slots else 0
    
    if scada_cache:
        print(f"\nProcessing {len(matches)} time range(s) with {total_slots} total time slots...")
    
    for idx, (row_num, row_data) in enumerate(matches, 1):
        if from_time_col and to_time_col and from_time_col <= len(row_data) and to_time_col <= len(row_data):
            from_time_val = row_data[from_time_col - 1] if from_time_col > 0 else None
            to_time_val = row_data[to_time_col - 1] if to_time_col > 0 else None
            date_val = row_data[date_col - 1] if date_col and date_col > 0 and date_col <= len(row_data) else None
            
            if from_time_val is not None and to_time_val is not None:
                slots = slots_15min(from_time_val, to_time_val)
                if slots:
                    date_str = format_value(date_val) if date_val else ""
                    
                    for slot_idx, (slot_from, slot_to) in enumerate(slots):
                        # Show progress for new dates
                        if date_str and date_str != current_date:
                            current_date = date_str
                            if scada_cache:
                                print(f"\n  Processing date: {date_str}...", flush=True)
                        
                        # Write date in first slot of each time range group
                        if slot_idx == 0 and date_str:
                            output_sheet.cell(row=row_idx, column=1).value = date_str
                        else:
                            output_sheet.cell(row=row_idx, column=1).value = ""  # Empty for subsequent slots in same range
                        
                        output_sheet.cell(row=row_idx, column=2).value = slot_from
                        output_sheet.cell(row=row_idx, column=3).value = slot_to
                        
                        # Lookup DC value if DC file is provided
                        dc_value = None
                        if dc_wb and date_str:
                            sheet_name = convert_date_to_sheet_format(date_str)
                            if sheet_name:
                                debug_lookup = args.verbose
                                dc_value = find_dc_value(dc_wb, sheet_name, slot_from, slot_to, debug=debug_lookup)
                                if dc_value is not None:
                                    dc_found_count += 1
                                else:
                                    dc_not_found_count += 1
                            elif args.verbose:
                                print(f"  Warning: Could not convert date '{date_str}' to sheet format", file=sys.stderr)
                        elif args.verbose and slot_idx == 0 and not dc_wb:
                            print(f"  Warning: DC workbook not available for lookup", file=sys.stderr)
                        
                        output_sheet.cell(row=row_idx, column=4).value = dc_value if dc_value is not None else ""
                        
                        # Lookup SCADA value using cache
                        scada_value = None
                        if scada_cache and date_str:
                            # Show progress only for first slot of each date (when loading file)
                            show_progress_now = (slot_idx == 0)
                            scada_value = find_scada_value(scada_cache, date_str, slot_from, debug=args.verbose, show_progress=show_progress_now)
                            if scada_value is not None:
                                scada_found_count += 1
                            else:
                                scada_not_found_count += 1
                            
                            # Increment counter and show progress
                            processed_slots += 1
                            if slot_idx == len(slots) - 1:  # Last slot of this range
                                print(f" ({processed_slots}/{total_slots} slots)", flush=True)
                            elif processed_slots % 50 == 0:
                                print(".", end="", flush=True)
                        
                        output_sheet.cell(row=row_idx, column=5).value = scada_value if scada_value is not None else ""
                        
                        # Apply borders
                        for col in range(1, 6):
                            output_sheet.cell(row=row_idx, column=col).border = thin_border
                        
                        row_idx += 1
    
    # Adjust column widths
    output_sheet.column_dimensions['A'].width = 15  # Date
    output_sheet.column_dimensions['B'].width = 10  # From
    output_sheet.column_dimensions['C'].width = 10  # To
    output_sheet.column_dimensions['D'].width = 12  # DC (MW)
    output_sheet.column_dimensions['E'].width = 25  # As per SLDC Scada in MW
    
    # Generate output filename with station name and timestamp (human-readable format with AM/PM)
    # Best practice: Use dashes for all separators (safe on all OS, readable)
    now = datetime.now()
    date_part = now.strftime("%d-%b-%Y")
    # Format time as "2-30-25-PM" (dashes instead of colons, no spaces, no leading zero on hour)
    hour = now.hour % 12
    if hour == 0:
        hour = 12
    time_part = f"{hour}-{now.minute:02d}-{now.second:02d}-{now.strftime('%p')}"
    timestamp = f"{date_part}_{time_part}"
    station_safe = args.station_name.replace(" ", "_").replace("/", "_")
    output_filename = f"{station_safe}_{timestamp}.xlsx"
    
    # Create output directory if it doesn't exist
    output_dir = xlsx_path.parent / "output"
    output_dir.mkdir(exist_ok=True)
    
    output_path = output_dir / output_filename
    
    output_wb.save(output_path)
    print(f"\nOutput file created: {output_path}")
    
    # Show summary only if there were issues
    if dc_wb:
        total_dc_lookups = dc_found_count + dc_not_found_count
        if total_dc_lookups > 0 and dc_found_count == 0:
            print(f"\nWarning: No DC values found ({dc_not_found_count} lookups). Use --verbose for details.", file=sys.stderr)
    
    if bd_folder and args.scada_column:
        total_scada_lookups = scada_found_count + scada_not_found_count
        if total_scada_lookups > 0 and scada_found_count == 0:
            print(f"\nWarning: No SCADA values found ({scada_not_found_count} lookups). Use --verbose for details.", file=sys.stderr)
    
    # Close all workbooks and caches
    wb.close()
    if dc_wb:
        dc_wb.close()
    if scada_cache:
        scada_cache.close_all()


if __name__ == "__main__":
    main()
