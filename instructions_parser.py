"""Extract station names and date range title from instructions Excel file."""

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

import find_station_rows as fsr

DATE_FORMATS = ["%d-%b-%Y", "%d-%b-%y", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"]
MAX_ROWS_TO_CHECK = 10_000
MAX_HEADER_COLS = 50


def _parse_dates_to_title(dates_found: list[str]) -> str:
    """Build report title string from list of date strings."""
    if not dates_found:
        return "⚡ GENERATE REPORT"
    parsed = []
    for d in dates_found:
        for fmt in DATE_FORMATS:
            try:
                parsed.append((datetime.strptime(d, fmt), d))
                break
            except ValueError:
                continue
    if parsed:
        parsed.sort(key=lambda x: x[0])
        from_d, to_d = parsed[0][1], parsed[-1][1]
        return f"⚡ GENERATE REPORT FROM {from_d} TO {to_d}" if from_d != to_d else f"⚡ GENERATE REPORT FROM {from_d}"
    dates_sorted = sorted(set(dates_found))
    if len(dates_sorted) == 1:
        return f"⚡ GENERATE REPORT FROM {dates_sorted[0]}"
    if len(dates_sorted) > 1:
        return f"⚡ GENERATE REPORT FROM {dates_sorted[0]} TO {dates_sorted[-1]}"
    return "⚡ GENERATE REPORT"


def extract_stations_and_title(
    file_path: Path,
    column_name: str,
    sheet_name: str = "",
) -> tuple[list[str], str]:
    """
    Read instructions Excel and return (station_names, report_title).
    Uses active sheet if sheet_name is empty.
    """
    station_names: list[str] = []
    report_title = "⚡ GENERATE REPORT"
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            target = sheet_name.strip().lower()
            sheet_found = None
            for name in wb.sheetnames:
                if name.strip().lower() == target or target in name.strip().lower():
                    sheet_found = name
                    break
            ws = wb[sheet_found] if sheet_found else wb.active
        else:
            ws = wb.active

        col_idx, header_row = fsr.find_column_by_name(ws, column_name, max_header_rows=10)
        if not col_idx:
            return [], report_title

        # Find date column
        date_col = None
        for c in range(1, min(ws.max_column + 1, MAX_HEADER_COLS)):
            val = ws.cell(row=header_row, column=c).value
            if val and "date" in str(val).strip().lower():
                date_col = c
                break

        data_start = (header_row or 1) + 1
        max_row = min(ws.max_row + 1, data_start + MAX_ROWS_TO_CHECK)
        unique_stations: set[str] = set()
        dates_found: list[str] = []

        for row_num in range(data_start, max_row):
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value:
                s = str(cell.value).strip()
                if s:
                    unique_stations.add(s)
            if date_col:
                date_cell = ws.cell(row=row_num, column=date_col)
                if date_cell.value:
                    date_val = fsr.format_value(date_cell.value)
                    if date_val:
                        dates_found.append(date_val)

        station_names = sorted(unique_stations)
        report_title = _parse_dates_to_title(dates_found)
    finally:
        wb.close()

    return station_names, report_title
