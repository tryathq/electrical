#!/usr/bin/env python3
"""
Streamlit Desktop App for Find Station Rows
Converts the command-line tool into a user-friendly GUI
"""

import json
import os
import sys
import tempfile
import threading
import uuid
from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st

from background_job import read_job as background_read_job
from background_job import write_job as background_write_job
from config import BACKGROUND_JOB_FILE, PAGE_SIZE_ALL, PARTIAL_OUTPUT_WRITE_INTERVAL, PROCESSING_BATCH_SIZE, REPORTS_DIR, table_height
from excel_builder import build_report_workbook
from instructions_parser import extract_stations_and_title
from reports_store import append_entry as reports_append_entry
from reports_store import load_index as reports_load_index
from reports_store import save_file as reports_save_file
from url_utils import url_main, url_report_file, url_reports_list

try:
    import openpyxl
except ImportError:
    st.error("❌ Missing dependency: openpyxl. Please install with: pip install openpyxl")
    st.stop()

# Try to import streamlit-aggrid for advanced table features
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode, JsCode
    AGGrid_AVAILABLE = True
except ImportError:
    AGGrid_AVAILABLE = False

# Add current directory to path to import find_station_rows module
sys.path.insert(0, str(Path(__file__).parent))

# Import the module (it will execute, but we'll use its functions)
try:
    import find_station_rows as fsr
    # Get the functions we need from find_station_rows
    format_value = fsr.format_value
    slots_15min = fsr.slots_15min
    time_to_minutes = fsr.time_to_minutes
    convert_date_to_sheet_format = fsr.convert_date_to_sheet_format
    SCADALookupCache = fsr.SCADALookupCache
    find_scada_value = fsr.find_scada_value
    find_dc_value = fsr.find_dc_value
    find_column_by_name = fsr.find_column_by_name
    find_matching_rows = fsr.find_matching_rows
except ImportError as e:
    st.error(f"Failed to import find_station_rows module: {e}")
    st.stop()


def _parse_float(val, default: float) -> float:
    """Parse value to float; return default if invalid or empty."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _reconstruct_ins_end_marker(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reconstruct the _ins_end marker for DataFrames loaded from Excel.
    If _ins_end column exists (from saved Excel), convert string values to boolean.
    Otherwise, try to detect instruction ends from the data pattern.
    """
    if "_ins_end" in df.columns:
        # Convert string values ("TRUE"/"FALSE") to boolean
        df = df.copy()
        df["_ins_end"] = df["_ins_end"].apply(
            lambda x: True if str(x).upper() == "TRUE" else False
        )
        return df
    
    df = df.copy()
    df["_ins_end"] = False
    
    if "From" not in df.columns or "To" not in df.columns or "Date" not in df.columns:
        return df
    
    n = len(df)
    for i in range(n):
        row = df.iloc[i]
        to_val = str(row.get("To", "")).strip() if pd.notna(row.get("To")) else ""
        from_val = str(row.get("From", "")).strip() if pd.notna(row.get("From")) else ""
        
        # Skip rows without From/To (Sum Mus rows)
        if not to_val or not from_val:
            continue
        
        # Check if next row starts a new instruction (has Date value)
        if i + 1 < n:
            next_row = df.iloc[i + 1]
            next_date = str(next_row.get("Date", "")).strip() if pd.notna(next_row.get("Date")) else ""
            next_from = str(next_row.get("From", "")).strip() if pd.notna(next_row.get("From")) else ""
            
            # Next row has a Date = new instruction starts
            # The row BEFORE next_date is either instruction end or gap end
            # We need to find the actual instruction end (where gap starts, if any)
            if next_date:
                # Look backwards from current row to find where instruction ends
                # Instruction end is where To time matches next_date row's From time
                # OR where there's a discontinuity (gap starts)
                
                # Walk backwards to find first row of current block (has Date or follows Sum Mus)
                block_start_idx = i
                for j in range(i, -1, -1):
                    check_row = df.iloc[j]
                    check_date = str(check_row.get("Date", "")).strip() if pd.notna(check_row.get("Date")) else ""
                    check_from = str(check_row.get("From", "")).strip() if pd.notna(check_row.get("From")) else ""
                    if check_date:
                        block_start_idx = j
                        break
                    if not check_from:  # Hit Sum Mus row
                        block_start_idx = j + 1
                        break
                
                # Now scan forward from block_start to find where To doesn't match next From
                # That's where instruction ends and gap begins
                found_gap_start = False
                for j in range(block_start_idx, i):
                    curr_to = str(df.iloc[j].get("To", "")).strip() if pd.notna(df.iloc[j].get("To")) else ""
                    next_f = str(df.iloc[j + 1].get("From", "")).strip() if pd.notna(df.iloc[j + 1].get("From")) else ""
                    if curr_to and next_f and curr_to != next_f:
                        # Gap found - row j is instruction end
                        df.iloc[j, df.columns.get_loc("_ins_end")] = True
                        found_gap_start = True
                        break
                
                # If no gap found, the row before next Date is instruction end
                if not found_gap_start:
                    df.iloc[i, df.columns.get_loc("_ins_end")] = True
                continue
            
            # Next row is Sum Mus row (no From value)
            if not next_from:
                # Similar logic: find where gap starts within this block
                block_start_idx = i
                for j in range(i, -1, -1):
                    check_row = df.iloc[j]
                    check_date = str(check_row.get("Date", "")).strip() if pd.notna(check_row.get("Date")) else ""
                    check_from = str(check_row.get("From", "")).strip() if pd.notna(check_row.get("From")) else ""
                    if check_date:
                        block_start_idx = j
                        break
                    if not check_from:
                        block_start_idx = j + 1
                        break
                
                found_gap_start = False
                for j in range(block_start_idx, i):
                    curr_to = str(df.iloc[j].get("To", "")).strip() if pd.notna(df.iloc[j].get("To")) else ""
                    next_f = str(df.iloc[j + 1].get("From", "")).strip() if pd.notna(df.iloc[j + 1].get("From")) else ""
                    if curr_to and next_f and curr_to != next_f:
                        df.iloc[j, df.columns.get_loc("_ins_end")] = True
                        found_gap_start = True
                        break
                
                if not found_gap_start:
                    df.iloc[i, df.columns.get_loc("_ins_end")] = True
        else:
            # Last row in dataframe
            if to_val:
                df.iloc[i, df.columns.get_loc("_ins_end")] = True
    
    return df


def _run_report_generation_worker(job_data: dict) -> None:
    """Run full report generation in a background thread. Updates job state file for progress."""
    temp_path = Path(job_data["temp_path"])
    instructions_name = job_data["instructions_name"]
    dc_name = job_data.get("dc_name") or ""
    bd_folder_path = job_data.get("bd_folder_path") or ""
    sheet_name = job_data.get("sheet_name") or ""
    column_name = job_data.get("column_name") or "Name of the station"
    station_name = job_data.get("station_name") or ""
    header_rows = int(job_data.get("header_rows", 10))
    data_only = bool(job_data.get("data_only", False))
    bd_sheet = job_data.get("bd_sheet") or ""
    scada_column = job_data.get("scada_column") or ""
    report_title = job_data.get("report_title") or "Back Down Calculator"
    ramp_up_5 = float(job_data.get("ramp_up_5", 15))
    ramp_up_10 = float(job_data.get("ramp_up_10", 27.5))
    ramp_up_15 = float(job_data.get("ramp_up_15", 40))
    ramp_down_5 = float(job_data.get("ramp_down_5", 15))
    ramp_down_10 = float(job_data.get("ramp_down_10", 27.5))
    ramp_down_15 = float(job_data.get("ramp_down_15", 40))
    verbose = False

    def update_progress(**kwargs):
        d = background_read_job() or {}
        d.update(kwargs)
        background_write_job(d)

    try:
        instructions_path = temp_path / instructions_name
        dc_path = temp_path / dc_name if dc_name and (temp_path / dc_name).exists() else None

        bd_folder = None
        if bd_folder_path:
            bd_folder = Path(bd_folder_path)
            if not bd_folder.exists():
                for pp in [Path(bd_folder_path), Path("data") / "BD", Path("data") / bd_folder_path]:
                    if pp.exists() and pp.is_dir():
                        bd_folder = pp
                        break
            if not bd_folder or not bd_folder.exists() or not bd_folder.is_dir():
                bd_folder = None

        wb = openpyxl.load_workbook(instructions_path, read_only=True, data_only=data_only)
        if sheet_name:
            sheet_found = None
            target = sheet_name.strip().lower()
            for name in wb.sheetnames:
                if name.strip().lower() == target or target in name.strip().lower():
                    sheet_found = name
                    break
            ws = wb[sheet_found] if sheet_found else wb.active
        else:
            ws = wb.active

        col_idx, header_row = find_column_by_name(ws, column_name, max_header_rows=header_rows)
        if col_idx is None:
            wb.close()
            update_progress(status="error", error_message=f"Column '{column_name}' not found")
            return

        matches = find_matching_rows(ws, col_idx, station_name, header_row)
        if not matches:
            wb.close()
            update_progress(status="error", error_message="No matching rows found")
            return

        from_time_col = to_time_col = date_col = None
        from_date_col = None  # Prefer "From Date" for instruction block date
        to_load_col = None  # To Load (MW) = floor for MW as per ramp
        for c in range(1, ws.max_column + 1):
            val = (ws.cell(row=header_row, column=c).value or "").strip().lower()
            if "from" in val and "time" in val:
                from_time_col = c
            elif "to" in val and "time" in val:
                to_time_col = c
            elif "from" in val and "date" in val:
                from_date_col = c
            elif "date" in val and date_col is None:
                date_col = c
            elif "to" in val and "load" in val:
                to_load_col = c
        if from_date_col is not None:
            date_col = from_date_col

        start_data_row = 2
        scada_cache = None
        if bd_folder and scada_column:
            scada_cache = SCADALookupCache(bd_folder, scada_column, bd_sheet if bd_sheet else None)
        dc_wb = openpyxl.load_workbook(dc_path, read_only=True, data_only=True) if dc_path else None

        total_slots = 0
        for _idx, (_row_num, row_data) in enumerate(matches, 1):
            if from_time_col and to_time_col and from_time_col <= len(row_data) and to_time_col <= len(row_data):
                from_time_val = row_data[from_time_col - 1] if from_time_col > 0 else None
                to_time_val = row_data[to_time_col - 1] if to_time_col > 0 else None
                if from_time_val is not None and to_time_val is not None:
                    slots = slots_15min(from_time_val, to_time_val)
                    total_slots += len(slots) if slots else 0

        output_rows = []
        dc_found_count = dc_not_found_count = scada_found_count = scada_not_found_count = 0
        processed_slots = 0
        current_date = None
        previous_date_with_data = None
        date_start_row = None
        row_idx = start_data_row
        last_progress_update = [0]
        entry_start_idx = 0  # Track start of current instruction entry
        pending_entry_start_idx = None  # Track start idx for pending Sum Mus calculation
        prev_instruction_end_time = None  # "HH:MM" of last slot To of previous instruction
        prev_instruction_end_mw_ramp = None  # last MW as per ramp of previous instruction
        prev_instruction_date_str = None  # date for gap rows between blocks

        def _num_display(val, decimals=2):
            """Format value for DC/SCADA columns: numeric to 2 decimals, else as-is."""
            if val is None or val == "":
                return ""
            try:
                n = float(val) if isinstance(val, (int, float, str)) and str(val).strip() else None
                return round(n, decimals) if n is not None else val
            except (ValueError, TypeError):
                return val

        for idx, (row_num, row_data) in enumerate(matches, 1):
            if not (from_time_col and to_time_col and from_time_col <= len(row_data) and to_time_col <= len(row_data)):
                continue
            from_time_val = row_data[from_time_col - 1] if from_time_col > 0 else None
            to_time_val = row_data[to_time_col - 1] if to_time_col > 0 else None
            date_val = row_data[date_col - 1] if date_col and date_col <= len(row_data) else None
            if from_time_val is None or to_time_val is None:
                continue
            slots = slots_15min(from_time_val, to_time_val)
            if not slots:
                continue
            date_str = format_value(date_val) if date_val else ""
            if date_str and date_str != previous_date_with_data and previous_date_with_data is not None:
                date_start_row = None
            if date_str and date_str != current_date:
                current_date = date_str
                previous_date_with_data = date_str
                date_start_row = row_idx

            # Check if there's a gap from previous instruction to this one
            # Gap info is stored and will be added AFTER the previous instruction's Sum Mus
            first_slot_from = slots[0][0] if slots else None
            there_was_gap = (
                prev_instruction_end_time is not None
                and first_slot_from is not None
                and str(prev_instruction_end_time).strip() != str(first_slot_from).strip()
                and prev_instruction_date_str
            )

            # Helper for gap processing
            def _time_to_minutes(t):
                try:
                    parts = str(t).strip().split(":")
                    return int(parts[0]) * 60 + int(parts[1])
                except:
                    return 0

            # If there was a pending instruction (not first iteration), add its gap rows and Sum Mus now
            if pending_entry_start_idx is not None:
                # Add gap rows from previous instruction end to current instruction start
                if there_was_gap:
                    gap_slots = slots_15min(prev_instruction_end_time, from_time_val)
                    gap_prev_mw = prev_instruction_end_mw_ramp
                    prev_end_mins = _time_to_minutes(prev_instruction_end_time)
                    dates_differ = (prev_instruction_date_str != date_str)
                    last_added_g_to = None
                    last_added_g_mw = None

                    for g_from, g_to in gap_slots:
                        g_from_mins = _time_to_minutes(g_from)
                        if dates_differ:
                            if g_from_mins < prev_end_mins:
                                gap_date_lookup = date_str
                            else:
                                gap_date_lookup = prev_instruction_date_str
                        else:
                            gap_date_lookup = prev_instruction_date_str

                        g_dc = None
                        if dc_wb and gap_date_lookup:
                            sheet_name_dc = convert_date_to_sheet_format(gap_date_lookup)
                            if sheet_name_dc:
                                g_dc = find_dc_value(dc_wb, sheet_name_dc, g_from, g_to, debug=verbose)
                        g_scada = None
                        if scada_cache and gap_date_lookup:
                            g_scada = find_scada_value(scada_cache, gap_date_lookup, g_from, debug=verbose, show_progress=False)
                        try:
                            g_dc_num = float(g_dc) if g_dc is not None else None
                        except (ValueError, TypeError):
                            g_dc_num = None
                        try:
                            g_scada_num = float(g_scada) if g_scada is not None else None
                        except (ValueError, TypeError):
                            g_scada_num = None

                        if gap_prev_mw is not None:
                            would_be = gap_prev_mw + ramp_up_15
                            if verbose:
                                print(f"  [GAP] {g_from}-{g_to}: prev_mw={gap_prev_mw:.2f}, would_be={would_be:.2f}, scada={g_scada_num}, dc={g_dc_num}", file=sys.stderr)
                            if g_scada_num is not None and would_be > g_scada_num:
                                if verbose:
                                    print(f"  [GAP] STOPPING: would_be {would_be:.2f} > scada {g_scada_num}", file=sys.stderr)
                                break
                            g_mw_ramp = would_be
                            if g_dc_num is not None and g_mw_ramp > g_dc_num:
                                g_mw_ramp = g_dc_num
                        else:
                            g_mw_ramp = None

                        gap_prev_mw = g_mw_ramp
                        # DC , Scada Diff (MW) = DC - Scada
                        g_diff = round(g_dc_num - g_scada_num, 2) if g_dc_num is not None and g_scada_num is not None else None
                        g_mus = round(g_diff / 4000, 10) if g_diff is not None else None
                        # Diff = Scada - MW as per ramp
                        g_scada_mw_diff = round(g_scada_num - g_mw_ramp, 2) if g_scada_num is not None and g_mw_ramp is not None else None
                        # MU = Diff/4000 if > 0, else 0
                        g_mu = round(g_scada_mw_diff / 4000, 10) if g_scada_mw_diff is not None and g_scada_mw_diff / 4000 > 0 else 0
                        # Gap rows have no Date (continue from previous instruction)
                        output_rows.append({
                            "Date": "",
                            "From": g_from,
                            "To": g_to,
                            "DC (MW)": _num_display(g_dc) if g_dc is not None else "",
                            "As per SLDC Scada in MW": _num_display(g_scada) if g_scada is not None else "",
                            "MW as per ramp": round(g_mw_ramp, 2) if g_mw_ramp is not None else "",
                            "DC , Scada Diff (MW)": g_diff if g_diff is not None else "",
                            "Mus": g_mus if g_mus is not None else "",
                            "Sum Mus": "",
                            "Diff": g_scada_mw_diff if g_scada_mw_diff is not None else "",
                            "MU": g_mu if g_mu is not None else "",
                            "Sum MU": "",
                            "_ins_end": False,  # Gap rows are not instruction ends
                        })
                        row_idx += 1
                        last_added_g_to = g_to
                        last_added_g_mw = g_mw_ramp

                    # Update prev values for continuity with next instruction (use last ADDED row's values)
                    if last_added_g_to is not None:
                        prev_instruction_end_time = last_added_g_to
                        prev_instruction_end_mw_ramp = last_added_g_mw

                # Now add Sum Mus for the previous instruction (including gap rows just added)
                entry_end_idx = len(output_rows)
                if entry_end_idx > pending_entry_start_idx:
                    mus_sum = 0.0
                    mu_sum = 0.0
                    for i in range(pending_entry_start_idx, entry_end_idx):
                        mus_val = output_rows[i].get("Mus")
                        if mus_val != "" and mus_val is not None:
                            try:
                                mus_sum += float(mus_val)
                            except (TypeError, ValueError):
                                pass
                        mu_val = output_rows[i].get("MU")
                        if mu_val != "" and mu_val is not None:
                            try:
                                mu_sum += float(mu_val)
                            except (TypeError, ValueError):
                                pass
                    mus_sum_rounded = round(mus_sum, 3) if mus_sum else 0.0
                    mu_sum_rounded = round(mu_sum, 3) if mu_sum else 0.0
                    output_rows.append({
                        "Date": "", "From": "", "To": "", "DC (MW)": "",
                        "As per SLDC Scada in MW": "", "MW as per ramp": "",
                        "DC , Scada Diff (MW)": "", "Mus": "", "Sum Mus": mus_sum_rounded, "Diff": "", "MU": "", "Sum MU": mu_sum_rounded,
                        "_ins_end": False,  # Sum rows are not instruction ends
                    })
                    row_idx += 1

            # To Load (floor for ramp down) from instruction row
            to_load = None
            if to_load_col and to_load_col <= len(row_data):
                try:
                    to_load = float(format_value(row_data[to_load_col - 1])) if row_data[to_load_col - 1] else None
                except (TypeError, ValueError):
                    to_load = None
            # Ramp down must not go below 270 (min floor); use To Load from row if higher
            floor_mw = max(270.0, to_load) if to_load is not None else 270.0
            prev_slot_mw_ramp = None

            # Start of this instruction's block (gap rows will be added after this instruction, before Sum Mus)
            entry_start_idx = len(output_rows)

            for slot_idx, (slot_from, slot_to) in enumerate(slots):
                # Show date at start of each instruction entry (first slot of this row only)
                row_date = date_str if (slot_idx == 0 and date_str) else ""
                dc_value = None
                if dc_wb and date_str:
                    sheet_name_dc = convert_date_to_sheet_format(date_str)
                    if sheet_name_dc:
                        dc_value = find_dc_value(dc_wb, sheet_name_dc, slot_from, slot_to, debug=verbose)
                        if dc_value is not None:
                            dc_found_count += 1
                        else:
                            dc_not_found_count += 1
                scada_value = None
                if scada_cache and date_str:
                    scada_value = find_scada_value(scada_cache, date_str, slot_from, debug=verbose, show_progress=False)
                    if scada_value is not None:
                        scada_found_count += 1
                    else:
                        scada_not_found_count += 1
                # MW as per ramp (rules from docs/MW_as_per_ramp_rules.md)
                dc_num = None
                if dc_value is not None:
                    try:
                        dc_num = float(dc_value) if isinstance(dc_value, (int, float, str)) and str(dc_value).strip() else None
                    except (ValueError, TypeError):
                        pass
                slot_min = time_to_minutes(slot_from)
                mw_as_per_ramp = None
                if slot_idx == 0:
                    # First slot of instruction block:
                    # - If continuous with prev (times match) OR gap was filled → use prev_mw - ramp_down
                    # - If no previous data → use DC - ramp_down (fresh start)
                    times_match = (
                        prev_instruction_end_time is not None
                        and str(slot_from).strip() == str(prev_instruction_end_time).strip()
                    )
                    # When gap rows were filled, prev_instruction_end_mw_ramp holds the last gap MW
                    # and prev_instruction_end_time equals slot_from, so times_match will be True
                    if times_match and prev_instruction_end_mw_ramp is not None:
                        # Continuous from previous (either direct or via filled gap rows)
                        # Apply ramp down from previous MW value
                        raw = prev_instruction_end_mw_ramp - ramp_down_15
                        try:
                            scada_num = float(scada_value) if scada_value is not None else None
                        except (ValueError, TypeError):
                            scada_num = None
                        if scada_num is not None and raw > scada_num:
                            raw = scada_num
                        mw_as_per_ramp = max(floor_mw, raw)
                        if prev_instruction_end_mw_ramp <= floor_mw:
                            mw_as_per_ramp = floor_mw
                    else:
                        # No previous data or times don't match → fresh start from DC - ramp_down
                        if prev_instruction_end_time is None:
                            gap_min = 15
                            if slot_min is not None:
                                gap_min = slot_min - (slot_min // 15) * 15
                                if gap_min == 0:
                                    gap_min = 15
                        else:
                            prev_min = time_to_minutes(prev_instruction_end_time)
                            if prev_min is not None and slot_min is not None:
                                gap_min = (slot_min - prev_min) % (24 * 60)
                                if gap_min <= 0:
                                    gap_min += 24 * 60
                            else:
                                gap_min = 15
                        ramp_down_val = ramp_down_15 if gap_min >= 15 else (ramp_down_10 if gap_min >= 10 else ramp_down_5)
                        mw_as_per_ramp = (dc_num - ramp_down_val) if dc_num is not None else None
                else:
                    # From second slot onward: always ramp down (continuous within block)
                    if prev_slot_mw_ramp is not None:
                        mw_as_per_ramp = max(floor_mw, prev_slot_mw_ramp - ramp_down_15)
                        if prev_slot_mw_ramp <= floor_mw:
                            mw_as_per_ramp = floor_mw
                    else:
                        mw_as_per_ramp = None
                prev_slot_mw_ramp = mw_as_per_ramp
                mw_ramp_display = round(mw_as_per_ramp, 2) if mw_as_per_ramp is not None else ""
                # Parse scada_num for calculations
                try:
                    scada_num = float(scada_value) if scada_value is not None else None
                except (ValueError, TypeError):
                    scada_num = None
                # DC , Scada Diff (MW) = DC - Scada
                diff_value = round(dc_num - scada_num, 2) if dc_num is not None and scada_num is not None else None
                mus_value = (float(diff_value) / 4000 if diff_value is not None else None) if diff_value is not None else None
                if diff_value is not None and mus_value is not None:
                    mus_value = round(mus_value, 10)
                # Diff = Scada - MW as per ramp
                scada_mw_diff = round(scada_num - mw_as_per_ramp, 2) if scada_num is not None and mw_as_per_ramp is not None else None
                # MU = Diff/4000 if > 0, else 0
                mu_value = round(scada_mw_diff / 4000, 10) if scada_mw_diff is not None and scada_mw_diff / 4000 > 0 else 0

                # Mark this row as instruction end if it's the last slot of the instruction
                is_instruction_end = (slot_to == slots[-1][1])
                
                output_rows.append({
                    "Date": row_date,
                    "From": slot_from,
                    "To": slot_to,
                    "DC (MW)": _num_display(dc_value) if dc_value is not None else "",
                    "As per SLDC Scada in MW": _num_display(scada_value) if scada_value is not None else "",
                    "MW as per ramp": mw_ramp_display,
                    "DC , Scada Diff (MW)": diff_value if diff_value is not None else "",
                    "Mus": mus_value if mus_value is not None else "",
                    "Sum Mus": "",
                    "Diff": scada_mw_diff if scada_mw_diff is not None else "",
                    "MU": mu_value if mu_value is not None else "",
                    "Sum MU": "",
                    "_ins_end": is_instruction_end,  # Hidden marker for styling
                })
                row_idx += 1
                processed_slots += 1
                if total_slots > 0 and processed_slots - last_progress_update[0] >= max(1, PROCESSING_BATCH_SIZE):
                    last_progress_update[0] = processed_slots
                    pct = min(99, int(100 * processed_slots / total_slots))
                    update_progress(processed_slots=processed_slots, total_slots=total_slots, progress_pct=pct, current_date=date_str or "")
                    # Write partial output every N slots to reduce I/O; also write first batch so table appears soon
                    if (
                        processed_slots % PARTIAL_OUTPUT_WRITE_INTERVAL == 0
                        or processed_slots == PROCESSING_BATCH_SIZE
                    ):
                        try:
                            partial_path = temp_path / "partial_output.json"
                            with open(partial_path, "w", encoding="utf-8") as f:
                                json.dump(output_rows, f, default=str, indent=0)
                        except Exception:
                            pass

            # End of this instruction: remember last slot for next block and date for gap rows
            if slots:
                prev_instruction_end_time = slots[-1][1]
                prev_instruction_end_mw_ramp = prev_slot_mw_ramp
                prev_instruction_date_str = date_str

            # Mark this instruction's start for deferred Sum Mus calculation (gap + Sum Mus added at start of next iteration)
            pending_entry_start_idx = entry_start_idx

        # After loop: add Sum Mus for the last instruction (no more instructions to trigger deferred processing)
        if pending_entry_start_idx is not None:
            entry_end_idx = len(output_rows)
            if entry_end_idx > pending_entry_start_idx:
                mus_sum = 0.0
                mu_sum = 0.0
                for i in range(pending_entry_start_idx, entry_end_idx):
                    mus_val = output_rows[i].get("Mus")
                    if mus_val != "" and mus_val is not None:
                        try:
                            mus_sum += float(mus_val)
                        except (TypeError, ValueError):
                            pass
                    mu_val = output_rows[i].get("MU")
                    if mu_val != "" and mu_val is not None:
                        try:
                            mu_sum += float(mu_val)
                        except (TypeError, ValueError):
                            pass
                mus_sum_rounded = round(mus_sum, 3) if mus_sum else 0.0
                mu_sum_rounded = round(mu_sum, 3) if mu_sum else 0.0
                output_rows.append({
                    "Date": "", "From": "", "To": "", "DC (MW)": "",
                    "As per SLDC Scada in MW": "", "MW as per ramp": "",
                    "DC , Scada Diff (MW)": "", "Mus": "", "Sum Mus": mus_sum_rounded, "Diff": "", "MU": "", "Sum MU": mu_sum_rounded,
                    "_ins_end": False,  # Sum rows are not instruction ends
                })

        wb.close()
        if dc_wb:
            dc_wb.close()
        if scada_cache:
            scada_cache.close_all()

        output_wb = build_report_workbook(output_rows)
        output_filename = f"{station_name.replace(' ', '_').replace('/', '_')}_{datetime.now().strftime('%d-%b-%Y_%H-%M-%S-%p')}.xlsx"
        output_path = temp_path / output_filename
        output_wb.save(output_path)

        reports_save_file(Path(output_path), output_filename)
        date_from = date_to = ""
        if " — " in report_title:
            part = report_title.split(" — ", 1)[1].strip()
            if " to " in part:
                date_from, date_to = (s.strip() for s in part.split(" to ", 1))
            else:
                date_from = part
        elif " FROM " in report_title.upper():
            # Parse "⚡ GENERATE REPORT FROM 01-Jan-2026 TO 31-Jan-2026" (from instructions_parser)
            idx_from = report_title.upper().index(" FROM ")
            part = report_title[idx_from + 6 :].strip()  # after " FROM "
            if " TO " in part.upper():
                idx_to = part.upper().index(" TO ")
                date_from = part[:idx_to].strip()
                date_to = part[idx_to + 4 :].strip()
            else:
                date_from = part
        if not date_from and output_rows:
            # Fallback: derive from actual data
            dates_in_data = [r.get("Date") for r in output_rows if r.get("Date")]
            if dates_in_data:
                date_from = min(dates_in_data)
                date_to = max(dates_in_data) if len(dates_in_data) > 1 else ""
        reports_append_entry({
            "filename": output_filename,
            "station": station_name,
            "date_from": date_from,
            "date_to": date_to,
            "run_at": datetime.now().isoformat(),
            "row_count": len(output_rows),
            "total_instructions": len(matches),
        })

        update_progress(
            status="done",
            output_filename=output_filename,
            progress_pct=100,
            processed_slots=processed_slots,
            total_slots=total_slots,
            total_instructions=len(matches),
            error_message=None,
        )
    except Exception as e:
        update_progress(status="error", error_message=str(e))


# Page config
st.set_page_config(
    page_title="Back Down Calculator",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Title will be updated after processing with date range
if 'report_title' not in st.session_state:
    st.session_state.report_title = "Back Down Calculator"
    st.session_state.report_subtitle = "Generate calculation sheet for BD and non compliance"

# Sync Reports view from URL (skip if user just navigated away so we don't re-apply stale URL)
if not st.session_state.pop("_url_go_main", None) and getattr(st, "query_params", None):
    qp = st.query_params
    if qp.get("view") == "report" and qp.get("file"):
        report_file = qp.get("file")
        for entry in reports_load_index():
            if entry.get("filename") == report_file:
                st.session_state["reports_view_filename"] = report_file
                st.session_state["reports_view_entry"] = entry
                st.session_state["reports_view_from_list"] = True
                st.session_state.pop("view_mode", None)
                break
    elif qp.get("view") == "reports" and not st.session_state.get("reports_view_filename"):
        st.session_state["view_mode"] = "reports"
if getattr(st, "query_params", None) and not st.session_state.get("view_mode") and not st.session_state.get("reports_view_filename"):
    if st.query_params.get("view") or st.query_params.get("file"):
        url_main()

# Read background job once per run (used by sidebar and main content)
_cached_bg_job = background_read_job()

# Sidebar: Menu at top (big square buttons); then Home (generate form) or Reports (list of reports)
with st.sidebar:
    st.markdown('<div data-app-menu-row style="display:none" aria-hidden="true"></div>', unsafe_allow_html=True)
    view_mode = st.session_state.get("view_mode", "")
    _sidebar_home = not view_mode and not st.session_state.get("reports_view_filename")
    _on_report = view_mode == "reports" or st.session_state.get("reports_view_filename") or st.session_state.get("reports_view_active")
    col_h, col_r = st.columns(2)
    with col_h:
        if st.button("🏠 Home", key="sidebar_home", type="primary" if _sidebar_home else "secondary", width='stretch'):
            for key in ("view_mode", "reports_view_filename", "reports_view_entry", "reports_view_active", "reports_view_from_list"):
                st.session_state.pop(key, None)
            st.session_state["_url_go_main"] = True
            url_main()
            st.rerun()
    with col_r:
        if st.button("📂 Reports", key="sidebar_reports", type="primary" if _on_report else "secondary", width='stretch'):
            st.session_state["view_mode"] = "reports"
            url_reports_list()
            st.rerun()
    st.divider()
    if _sidebar_home:
        st.caption("**Home** — generate report")
        st.header("📋 Input Files")
    
        # Instructions file upload
        instructions_file = st.file_uploader(
        "Instructions Excel File",
        type=['xlsx', 'xls'],
        help="Upload the instructions XLSX file",
        key="instructions_file_upload"
        )
    
        # Sheet name (optional) - removed from UI, defaults to active sheet
        sheet_name = ""
    
        # Column name (read-only)
        column_name = st.text_input(
        "Column Name",
        value="Name of the station",
        help="Column header to search for station name",
        disabled=True
        )
    
        # Extract unique station names from file
        station_names = []
        station_name = None
    
        if instructions_file is not None:
            # Use session state to cache station names per file
            file_key = f"{instructions_file.name}_{sheet_name}_{column_name}"
            
            if 'station_names_cache' not in st.session_state:
                st.session_state.station_names_cache = {}
    
            # Always extract dates for title, even if station names are cached
            date_cache_key = f"{instructions_file.name}_{sheet_name}_dates"
            if 'date_range_cache' not in st.session_state:
                st.session_state.date_range_cache = {}
    
            if date_cache_key not in st.session_state.date_range_cache or file_key not in st.session_state.station_names_cache:
                with st.spinner("Extracting station names and dates from file..."):
                    tmp_path = None
                    try:
                        instructions_file.seek(0)
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                            tmp_file.write(instructions_file.getbuffer())
                            tmp_path = Path(tmp_file.name)
                        station_names, title_str = extract_stations_and_title(tmp_path, column_name, sheet_name)
                        st.session_state.station_names_cache[file_key] = station_names
                        st.session_state.date_range_cache[date_cache_key] = title_str
                        st.session_state.report_title = title_str
                    except Exception as e:
                        st.warning(f"Could not extract station names: {e}")
                        st.session_state.station_names_cache[file_key] = []
                        st.session_state.date_range_cache[date_cache_key] = "Back Down Calculator"
                        st.session_state.report_title = "Back Down Calculator"
                    finally:
                        if tmp_path and tmp_path.exists():
                            try:
                                tmp_path.unlink()
                            except Exception:
                                pass
            else:
                station_names = st.session_state.station_names_cache[file_key]
                # Restore title from cache
                if date_cache_key in st.session_state.date_range_cache:
                    st.session_state.report_title = st.session_state.date_range_cache[date_cache_key]
        
            # Show dropdown (always selectbox, never editable text input)
            if station_names:
                station_name = st.selectbox(
                    "Station Name",
                    options=station_names,
                    help=f"Select station name from the dropdown ({len(station_names)} stations found in file)",
                    key="station_selectbox"
                )
                st.caption(f"✓ Found {len(station_names)} unique station(s)")
            else:
                # Show empty selectbox (not editable) if no stations found
                station_name = st.selectbox(
                    "Station Name",
                    options=[],
                    help="No stations found in file. Please check the file and column name.",
                    disabled=True,
                    key="station_selectbox_empty"
                )
                if column_name:
                    st.caption("⚠️ No stations found. Check if column name matches the file.")
        else:
            # No file uploaded yet, show text input
            station_name = st.text_input(
                "Station Name",
                value="",
                help="Upload instructions file to see dropdown, or enter station name manually"
            )
        
        st.divider()
        st.header("⚙️ Options")
        
        # DC file upload (mandatory)
        dc_file = st.file_uploader(
            "DC File",
            type=['xlsx', 'xls'],
            help="DC Excel file with date-named sheets (required)"
        )
        
        # BD folder (mandatory)
        bd_folder_path = st.text_input(
            "BD Folder Path",
            value="",
            help="Path to folder containing SCADA BD files (required)"
        )
        
        # BD sheet name (mandatory) - extract from BD file
        bd_sheet = ""
        bd_sheet_options = []
        
        if bd_folder_path and bd_folder_path.strip():
            bd_folder = Path(bd_folder_path.strip())
            if bd_folder.exists() and bd_folder.is_dir():
                # Find first Excel file in BD folder
                bd_files = list(bd_folder.glob("*.xlsx")) + list(bd_folder.glob("*.xls"))
                if bd_files:
                    bd_file_path = bd_files[0]
                    file_key_sheets = f"{bd_file_path.name}_sheets"
                
                    if 'bd_sheets_cache' not in st.session_state:
                        st.session_state.bd_sheets_cache = {}
                
                    if file_key_sheets not in st.session_state.bd_sheets_cache:
                        try:
                            with st.spinner("Extracting sheet names from BD file..."):
                                wb_bd = openpyxl.load_workbook(bd_file_path, read_only=True, data_only=True)
                                bd_sheet_options = wb_bd.sheetnames
                                st.session_state.bd_sheets_cache[file_key_sheets] = bd_sheet_options
                                wb_bd.close()
                        except Exception as e:
                            st.session_state.bd_sheets_cache[file_key_sheets] = []
                    else:
                        bd_sheet_options = st.session_state.bd_sheets_cache[file_key_sheets]
        
        # Show BD sheet dropdown if options available
        if bd_sheet_options:
            bd_sheet = st.selectbox(
                "BD Sheet Name",
                options=bd_sheet_options,
                help="Select sheet name from BD file (extracted from BD folder)",
                key="bd_sheet_selectbox"
            )
            st.caption(f"✓ Found {len(bd_sheet_options)} sheet(s) in BD file")
        else:
            bd_sheet = st.text_input(
                "BD Sheet Name",
                value="",
                help="Sheet name in BD files (e.g., 'DATA-CMD') (required)",
                key="bd_sheet_text"
            )
            if bd_folder_path:
                st.caption("⚠️ Could not extract sheets. Check BD folder path.")
        
        # SCADA column (mandatory) - extract from BD file
        scada_column = None
        scada_column_options = []
        
        if bd_folder_path and bd_folder_path.strip() and bd_sheet and str(bd_sheet).strip():
            bd_folder = Path(bd_folder_path.strip())
            if bd_folder.exists() and bd_folder.is_dir():
                # Find first Excel file in BD folder
                bd_files = list(bd_folder.glob("*.xlsx")) + list(bd_folder.glob("*.xls"))
                if bd_files:
                    bd_file_path = bd_files[0]
                    file_key_cols = f"{bd_file_path.name}_{bd_sheet.strip()}_columns"
                
                    if 'bd_columns_cache' not in st.session_state:
                        st.session_state.bd_columns_cache = {}
                
                    if file_key_cols not in st.session_state.bd_columns_cache:
                        try:
                            with st.spinner("Extracting column names from BD file..."):
                                wb_bd = openpyxl.load_workbook(bd_file_path, read_only=True, data_only=True)
                            
                                # Find the specified sheet
                                sheet_found = None
                                target_sheet = bd_sheet.strip().lower()
                                for name in wb_bd.sheetnames:
                                    if name.strip().lower() == target_sheet or target_sheet in name.strip().lower():
                                        sheet_found = name
                                        break
                            
                                if sheet_found:
                                    ws_bd = wb_bd[sheet_found]
                                    # Extract column names only from header row (typically row 1)
                                    column_names = []
                                    # Most Excel files have headers in row 1
                                    header_row = 1
                                
                                    # Extract all values from header row
                                    for col_idx in range(1, min(ws_bd.max_column + 1, 200)):
                                        cell = ws_bd.cell(row=header_row, column=col_idx)
                                        if cell.value:
                                            col_name = str(cell.value).strip()
                                            if col_name:
                                                column_names.append(col_name)
                                
                                    # If row 1 is empty or has very few values, try row 2
                                    if len(column_names) < 2 and ws_bd.max_row >= 2:
                                        column_names = []
                                        header_row = 2
                                        for col_idx in range(1, min(ws_bd.max_column + 1, 200)):
                                            cell = ws_bd.cell(row=header_row, column=col_idx)
                                            if cell.value:
                                                col_name = str(cell.value).strip()
                                                if col_name:
                                                    column_names.append(col_name)
                                
                                    scada_column_options = sorted(list(set(column_names)))  # Remove duplicates, sort
                                    st.session_state.bd_columns_cache[file_key_cols] = scada_column_options
                                else:
                                    st.session_state.bd_columns_cache[file_key_cols] = []
                            
                                wb_bd.close()
                        except Exception as e:
                            st.session_state.bd_columns_cache[file_key_cols] = []
                    else:
                        scada_column_options = st.session_state.bd_columns_cache[file_key_cols]
        
        # Show SCADA column dropdown if options available
        if scada_column_options:
            scada_column = st.selectbox(
                "SCADA Column Name",
                options=scada_column_options,
                help="Select column name from BD file (extracted from BD folder)",
                key="scada_column_selectbox"
            )
            st.caption(f"✓ Found {len(scada_column_options)} column(s) in BD file")
        else:
            scada_column = st.text_input(
                "SCADA Column Name",
                value="",
                help="Column header name in BD files (e.g., 'HNJA4_AG.STTN.X_BUS_GEN.MW') (required)",
                key="scada_column_text"
            )
            if bd_folder_path and bd_sheet:
                st.caption("⚠️ Could not extract columns. Check BD folder path and sheet name.")
        
        st.divider()
        st.header("📈 Ramp Rates")
        st.caption("**Ramp Up** (MW)", help="Ramp up rate in MW for 5, 10, 15 min gaps")
        ru1, ru2, ru3 = st.columns(3)
        with ru1:
            ramp_up_5 = st.text_input("5 min", value="15", placeholder="15", key="ramp_up_5_input")
        with ru2:
            ramp_up_10 = st.text_input("10 min", value="27.5", placeholder="27.5", key="ramp_up_10_input")
        with ru3:
            ramp_up_15 = st.text_input("15 min", value="40", placeholder="40", key="ramp_up_15_input")
        st.caption("**Ramp Down** (MW)", help="Ramp down rate in MW for 5, 10, 15 min gaps")
        rd1, rd2, rd3 = st.columns(3)
        with rd1:
            ramp_down_5 = st.text_input("5 min", value="15", placeholder="15", key="ramp_down_5_input")
        with rd2:
            ramp_down_10 = st.text_input("10 min", value="27.5", placeholder="27.5", key="ramp_down_10_input")
        with rd3:
            ramp_down_15 = st.text_input("15 min", value="40", placeholder="40", key="ramp_down_15_input")
        
        # Defaults (advanced options removed for now)
        header_rows = 10
        data_only = False
        verbose = False
        
        # Generate button at bottom of sidebar - enabled only when all required fields are filled
        st.divider()
        _bg_status_sidebar = _cached_bg_job.get("status") if _cached_bg_job else None
        
        # Check all required fields
        _all_fields_filled = (
            instructions_file is not None
            and station_name and str(station_name).strip()
            and dc_file is not None
            and bd_folder_path and str(bd_folder_path).strip()
            and scada_column and str(scada_column).strip()
            and bd_sheet and str(bd_sheet).strip()
        )
        
        if _bg_status_sidebar == "running":
            st.button("⏳ Generating...", type="secondary", use_container_width=True, disabled=True, key="sidebar_generate_disabled")
        elif not _all_fields_filled:
            st.button("🚀 Generate Report", type="primary", use_container_width=True, disabled=True, key="sidebar_generate_btn_disabled")
            # Show which fields are missing
            _missing = []
            if instructions_file is None:
                _missing.append("Instructions File")
            if not station_name or not str(station_name).strip():
                _missing.append("Station Name")
            if dc_file is None:
                _missing.append("DC File")
            if not bd_folder_path or not str(bd_folder_path).strip():
                _missing.append("BD Folder Path")
            if not bd_sheet or not str(bd_sheet).strip():
                _missing.append("BD Sheet Name")
            if not scada_column or not str(scada_column).strip():
                _missing.append("SCADA Column")
            if _missing:
                st.caption(f"⚠️ Missing: {', '.join(_missing)}")
        else:
            if st.button("🚀 Generate Report", type="primary", use_container_width=True, key="sidebar_generate_btn"):
                st.session_state["_sidebar_generate_clicked"] = True
                st.rerun()
    else:
        # Reports: show list of saved reports in sidebar; selecting one shows it on the right
        instructions_file = None
        station_name = ""
        dc_file = None
        bd_folder_path = ""
        scada_column = None
        bd_sheet = ""
        header_rows = 10
        data_only = False
        verbose = False
        st.caption("**Back Down reports** — select a report")
        reports_list_sidebar = reports_load_index()
        # Prepend in-progress report to list when a background job is running
        _bg_job_sidebar = _cached_bg_job
        if _bg_job_sidebar and _bg_job_sidebar.get("status") == "running":
            _gen_station = _bg_job_sidebar.get("station_name", "Report")
            _gen_entry = {
                "filename": "__generating__",
                "station": _gen_station,
                "date_from": "",
                "date_to": "",
                "run_at": _bg_job_sidebar.get("created_at", ""),
                "row_count": 0,
                "_generating": True,
            }
            reports_list_sidebar.insert(0, _gen_entry)
        if not reports_list_sidebar:
            st.info("No saved reports yet. Go to **Home** to generate one.")
        else:
            st.caption(f"{len(reports_list_sidebar)} report(s)")
            _selected_report = st.session_state.get("reports_view_filename") or st.session_state.get("reports_view_active")
            for i, entry in enumerate(reports_list_sidebar):
                fn = entry.get("filename", "")
                is_generating = fn == "__generating__" or entry.get("_generating")
                station = entry.get("station", "")
                date_from = entry.get("date_from", "")
                date_to = entry.get("date_to", "")
                if is_generating:
                    date_range = "generating…"
                    label = f"⏳ {station} — generating…"
                    generated_str = "In progress"
                else:
                    date_range = f"{date_from} → {date_to}" if date_to else (date_from or "—")
                    label = f"{station} — {date_range}"
                    run_at = entry.get("run_at", "")
                    try:
                        dt = datetime.fromisoformat(run_at.replace("Z", "+00:00"))
                        generated_str = dt.strftime("%d %b %Y, %I:%M %p")
                    except Exception:
                        generated_str = run_at if run_at else "—"
                label_with_time = f"{label}\n{generated_str}"
                _is_selected = (fn == _selected_report)
                c1, c2 = st.columns([5, 1])
                with c1:
                    if st.button(label_with_time, key=f"sidebar_rep_{i}_{fn}", type="primary" if _is_selected else "secondary", width='stretch', help="Show this report on the right" if not is_generating else "Show current progress"):
                        st.session_state["reports_view_filename"] = fn
                        st.session_state["reports_view_entry"] = entry
                        st.session_state["reports_view_from_list"] = True
                        url_report_file(fn)
                        st.rerun()
                with c2:
                    if not is_generating:
                        report_path = REPORTS_DIR / fn
                        if report_path.exists():
                            with open(report_path, "rb") as f:
                                st.download_button("📥", data=f.read(), file_name=fn, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"sidebar_dl_{i}_{fn}")

# Global CSS for sidebar menu buttons (square box look); report list: two-line label
st.markdown("""
<style>
    [data-app-menu-row] ~ [data-testid="stHorizontalBlock"] button,
    [data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:first-of-type button {
        font-size: 1.05rem !important;
        padding: 0.7rem !important;
        font-weight: 600 !important;
        min-height: 3rem !important;
        border-radius: 10px !important;
        border: 1px solid rgba(49, 51, 63, 0.2) !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
    }
    [data-app-menu-row] ~ [data-testid="stHorizontalBlock"] button {
        white-space: pre-line !important;
    }
</style>
""", unsafe_allow_html=True)

# Display title and subtitle per page (Reports list vs viewing a report vs Home)
_on_reports_list = st.session_state.get("view_mode") == "reports" and not st.session_state.get("reports_view_filename")
_viewing_saved_report_header = bool(st.session_state.get("reports_view_filename") and st.session_state.get("reports_view_entry"))
if _on_reports_list:
    title_to_show = "Back Down reports"
    subtitle_to_show = "Choose a report to view."
elif _viewing_saved_report_header:
    title_to_show = st.session_state.get('report_title', "Back Down Report")
    subtitle_to_show = "Generated calculation sheet for BD and non compliance"
else:
    title_to_show = st.session_state.get('report_title', "Back Down Calculator")
    subtitle_to_show = st.session_state.get('report_subtitle', "Generate calculation sheet for BD and non compliance")
st.title(title_to_show)
st.markdown(subtitle_to_show)

# Background report generation: show status on any page so user can navigate away
_bg_job = _cached_bg_job
_status = _bg_job.get("status") if _bg_job else None
_reports_view_filename = st.session_state.get("reports_view_filename")
_reports_view_entry = st.session_state.get("reports_view_entry")
_viewing_saved_report = bool(_reports_view_filename and _reports_view_entry)
# When job completes, clear "viewing generating report" so user is not stuck
if _status == "done" and _reports_view_filename == "__generating__":
    for key in ("reports_view_filename", "reports_view_entry", "reports_view_active", "reports_view_from_list"):
        st.session_state.pop(key, None)
    _reports_view_filename = _reports_view_entry = None
    _viewing_saved_report = False
if _status in ("running", "done", "error"):
    if _status == "running":
        # Only show "generating" banner when viewing Home or the in-progress report, not when viewing a completed report
        if not _viewing_saved_report or _reports_view_filename == "__generating__":
            _pct = _bg_job.get("progress_pct", 0)
            _processed = _bg_job.get("processed_slots", 0)
            _total = _bg_job.get("total_slots", 0) or 1
            _date = _bg_job.get("current_date", "")
            st.info(f"⏳ **Report generating in background** — {_pct}% ({_processed} / {_total} slots)" + (f" — {_date}" if _date else "") + ". You can switch to **Reports** or other pages; generation will continue.")
            if st.button("🔄 Refresh status", key="bg_job_refresh"):
                st.rerun()
    elif _status == "done":
        # Only show "Report ready" banner on Home page, not when viewing Reports page
        if not _viewing_saved_report and not _on_reports_list:
            st.success(f"✅ **Report ready.** Displaying below. Also saved to **Reports**.")
            # Automatically load and display the completed report on home page
            _done_filename = _bg_job.get("output_filename")
            if _done_filename:
                _done_report_path = REPORTS_DIR / _done_filename
                if _done_report_path.exists():
                    try:
                        _done_df = pd.read_excel(_done_report_path, engine="openpyxl")
                        _done_df = _reconstruct_ins_end_marker(_done_df)
                        _done_report_key = f"output_data_home_{_done_filename}"
                        st.session_state[_done_report_key] = _done_df
                        st.session_state["display_output_data_key"] = _done_report_key
                        st.session_state["display_station_name"] = _bg_job.get("station_name", "")
                        # Calculate stats
                        _done_total_days = 0
                        _done_total_instructions = _bg_job.get("total_instructions", 0)
                        if "Date" in _done_df.columns:
                            _done_unique_dates = _done_df["Date"].dropna().astype(str).replace("", pd.NA).dropna().unique()
                            _done_total_days = len([d for d in _done_unique_dates if d and d.strip()])
                        if not _done_total_instructions and "Sum Mus" in _done_df.columns:
                            _done_total_instructions = len(_done_df[_done_df["Sum Mus"].notna() & (_done_df["Sum Mus"] != "")])
                        st.session_state["display_stats"] = {
                            "total_days": _done_total_days,
                            "total_instructions": _done_total_instructions,
                            "output_rows": len(_done_df),
                        }
                        # Clear background job after loading
                        background_write_job({})
                    except Exception:
                        pass
    elif _status == "error":
        _err = _bg_job.get("error_message", "Unknown error")
        st.error(f"❌ **Report generation failed:** {_err}")
        if st.button("Dismiss", key="bg_job_error_dismiss"):
            background_write_job({})
            st.rerun()

# When Reports is selected but no report chosen yet: show prompt only (no duplicate header)
if _on_reports_list:
    url_reports_list()
    st.info("Select a report from the list on the left to view it here.")
    st.stop()

# Main content area (skip input checks when viewing a saved report from Reports list)
_reports_view_filename = st.session_state.get("reports_view_filename")
_reports_view_entry = st.session_state.get("reports_view_entry")
_viewing_saved_report = bool(_reports_view_filename and _reports_view_entry)

# On Home or when "generating" report selected from list: show live table view while background report is generating
_viewing_generating_report = _viewing_saved_report and _reports_view_filename == "__generating__"
if _status == "running" and _bg_job and (not _viewing_saved_report or _viewing_generating_report):
    _temp_path = Path(_bg_job.get("temp_path", ""))
    _partial_file = _temp_path / "partial_output.json" if _temp_path else None
    if _partial_file and _partial_file.exists():
        try:
            with open(_partial_file, "r", encoding="utf-8") as f:
                _partial_rows = json.load(f)
        except Exception:
            _partial_rows = []
        if not _partial_rows and _viewing_generating_report:
            st.caption("⏳ Waiting for first batch of data… Click **Refresh status** below to update.")
            if st.button("🔄 Refresh status", key="bg_job_refresh_wait"):
                st.rerun()
        elif _partial_rows:
            _total_slots = _bg_job.get("total_slots", 0) or 1
            _processed = _bg_job.get("processed_slots", 0)
            _pct = _bg_job.get("progress_pct", 0)
            _current_date = _bg_job.get("current_date", "")
            _station_bg = _bg_job.get("station_name", "")
            st.progress(_pct / 100.0)
            if _current_date:
                st.caption(f"⏳ Processing day {_current_date} — {len(_partial_rows)} rows so far")
            else:
                st.caption(f"⏳ Processing... {len(_partial_rows)} rows so far")
            _df_partial = pd.DataFrame(_partial_rows).fillna("").replace("None", "")
            for _col in ("DC (MW)", "As per SLDC Scada in MW", "MW as per ramp", "DC , Scada Diff (MW)", "Mus", "Sum Mus", "Diff", "MU", "Sum MU"):
                if _col in _df_partial.columns:
                    _df_partial[_col] = pd.to_numeric(_df_partial[_col], errors="coerce")
            if "DC , Scada Diff (MW)" in _df_partial.columns:
                _df_partial["DC , Scada Diff (MW)"] = _df_partial["DC , Scada Diff (MW)"].apply(
                    lambda x: round(x, 2) if isinstance(x, (int, float)) and pd.notna(x) else x
                )
            if "Sum Mus" in _df_partial.columns:
                _df_partial["Sum Mus"] = _df_partial["Sum Mus"].apply(
                    lambda x: round(x, 3) if isinstance(x, (int, float)) and pd.notna(x) else x
                )
            # Reorder columns to match expected output format (including hidden marker columns)
            _expected_cols = ["Date", "From", "To", "DC (MW)", "As per SLDC Scada in MW", "DC , Scada Diff (MW)", "Mus", "Sum Mus", "MW as per ramp", "Diff", "MU", "Sum MU", "_ins_end"]
            _df_partial = _df_partial[[c for c in _expected_cols if c in _df_partial.columns]]
            _title_parts = ["Calculation sheet for BD and non compliance of", _station_bg or "…"]
            st.divider()
            st.header(f"📊 {' '.join(_title_parts)} — ⏳ generating…")
            if AGGrid_AVAILABLE:
                _n_partial = len(_df_partial)
                _gb = GridOptionsBuilder.from_dataframe(_df_partial)
                _page_opts = sorted(set([20, 50, 100, 500, _n_partial])) if _n_partial > 0 else [20]
                _default_ps = _n_partial if _n_partial > 0 else 20
                _gb.configure_pagination(
                    paginationAutoPageSize=False,
                    paginationPageSize=_default_ps,
                )
                _gb.configure_grid_options(
                    paginationPageSizeSelector=_page_opts,
                    onFirstDataRendered=JsCode(
                        f"""
                        function(params) {{
                            var allVals = ['{PAGE_SIZE_ALL}', '{_n_partial}'];
                            function replacePageSizeText(root) {{
                                try {{
                                    var walk = document.createTreeWalker(root, NodeFilter.SHOW_TEXT);
                                    var n;
                                    while ((n = walk.nextNode())) {{
                                        var t = n.textContent.trim();
                                        if (allVals.indexOf(t) !== -1) n.textContent = 'ALL';
                                    }}
                                    if (root.querySelectorAll) root.querySelectorAll('select option').forEach(function(opt) {{
                                        if (allVals.indexOf(opt.value) !== -1) opt.textContent = 'ALL';
                                    }});
                                }} catch (e) {{}}
                            }}
                            function run() {{
                                var el = params.api.getGridElement();
                                if (el) {{
                                    var root = el.closest('.ag-root-wrapper') || el.closest('.ag-root') || el;
                                    if (root) replacePageSizeText(root);
                                }}
                                replacePageSizeText(document.body);
                            }}
                            setTimeout(run, 100);
                            setTimeout(run, 500);
                        }}
                        """
                    ),
                )
                _gb.configure_side_bar()
                _gb.configure_default_column(sortable=True, filterable=True, resizable=True, editable=False)
                _gb.configure_selection("single")
                
                # Cell styling for partial view
                _date_style = JsCode("""
                function(params) {
                    if (params.value && params.value.toString().trim() !== '') {
                        return {'backgroundColor': '#FFFF00', 'fontWeight': 'bold'};
                    }
                    return null;
                }
                """)
                _gb.configure_column("Date", cellStyle=_date_style)
                
                _to_style = JsCode("""
                function(params) {
                    var rowData = params.data;
                    var insEnd = rowData['_ins_end'];
                    if (insEnd === true || insEnd === 1 || insEnd === 'True' || insEnd === 'true' || insEnd === 'TRUE') {
                        return {'backgroundColor': '#FFFF00', 'fontWeight': 'bold'};
                    }
                    return null;
                }
                """)
                _gb.configure_column("To", cellStyle=_to_style)
                _gb.configure_column("_ins_end", hide=True)
                
                AgGrid(
                    _df_partial,
                    gridOptions=_gb.build(),
                    height=table_height(min(_n_partial, 100) if _n_partial > 0 else 20),
                    width="100%",
                    theme="streamlit",
                    update_mode=GridUpdateMode.NO_UPDATE,
                    data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                    allow_unsafe_jscode=True,
                )
            else:
                st.dataframe(_df_partial, width="stretch", height=table_height(len(_df_partial)), hide_index=True)
            if st.button("🔄 Refresh status", key="bg_job_refresh_table"):
                st.rerun()
# Show upload/form prompts only when not viewing a report and not in the middle of background generation
# But don't stop if we have a latest report to show
_has_reports_to_show = bool(reports_load_index())
if not _viewing_saved_report and _status != "running":
    if instructions_file is None:
        if not _has_reports_to_show:
            st.info("👈 Please upload an Instructions Excel file in the sidebar to get started.")
            st.stop()

    if not station_name or station_name.strip() == "":
        if not _has_reports_to_show:
            if 'station_names_cache' in st.session_state and len(st.session_state.station_names_cache) > 0:
                st.warning("⚠️ Please select a Station Name from the dropdown")
            else:
                st.warning("⚠️ Please enter or select a Station Name")
            st.stop()

    if dc_file is None:
        if not _has_reports_to_show:
            st.info("👈 Please upload a DC Excel file in the sidebar to get started.")
            st.stop()

    if not bd_folder_path or bd_folder_path.strip() == "":
        if not _has_reports_to_show:
            st.info("👈 Please enter the path to the BD folder in the sidebar.")
            st.stop()

    if not scada_column or scada_column.strip() == "":
        if not _has_reports_to_show:
            st.error("❌ SCADA Column Name is required. Please enter the SCADA column name.")
            st.stop()

    if not bd_sheet or bd_sheet.strip() == "":
        if not _has_reports_to_show:
            st.error("❌ BD Sheet Name is required. Please enter the BD sheet name.")
            st.stop()

# When viewing a past report from Menu Reports: load it into session and set display keys (once)
if _reports_view_filename and _reports_view_entry and st.session_state.get("reports_view_active") != _reports_view_filename:
    if _reports_view_filename == "__generating__":
        # In-progress report: no file to load; just set active and title from job
        st.session_state["reports_view_active"] = "__generating__"
        if _bg_job:
            st.session_state["report_title"] = _bg_job.get("report_title", "Back Down Report")
    else:
        report_key = f"output_data_report_{_reports_view_filename}"
        report_path = REPORTS_DIR / _reports_view_filename
        if report_path.exists():
            try:
                df_report = pd.read_excel(report_path, engine="openpyxl")
                df_report = _reconstruct_ins_end_marker(df_report)
                st.session_state[report_key] = df_report
                st.session_state["display_output_data_key"] = report_key
                st.session_state["display_station_name"] = _reports_view_entry.get("station", "")
                date_f = _reports_view_entry.get("date_from", "")
                date_t = _reports_view_entry.get("date_to", "")
                # Calculate total_days from unique dates in the data
                _total_days = 0
                _total_instructions = _reports_view_entry.get("total_instructions", 0)
                if "Date" in df_report.columns:
                    _unique_dates = df_report["Date"].dropna().astype(str).replace("", pd.NA).dropna().unique()
                    _total_days = len([d for d in _unique_dates if d and d.strip()])
                # If total_instructions not stored, count from Sum Mus rows
                if not _total_instructions and "Sum Mus" in df_report.columns:
                    _total_instructions = len(df_report[df_report["Sum Mus"].notna() & (df_report["Sum Mus"] != "")])
                st.session_state["display_stats"] = {
                    "total_days": _total_days,
                    "total_instructions": _total_instructions,
                    "output_rows": _reports_view_entry.get("row_count", 0) or len(df_report),
                }
                if date_f and date_t:
                    st.session_state["report_title"] = f"Back Down Report — {date_f} to {date_t}"
                elif date_f:
                    st.session_state["report_title"] = f"Back Down Report — {date_f}"
                else:
                    st.session_state["report_title"] = "Back Down Report"
                st.session_state["reports_view_active"] = _reports_view_filename
            except Exception:
                st.session_state["reports_view_active"] = None
        else:
            st.session_state["reports_view_active"] = None

# Display output data BEFORE processing - skip when we're showing background job live table (Home or Reports list)
_showing_bg_job_table = (
    _status == "running"
    and _bg_job
    and (Path(_bg_job.get("temp_path", "")) / "partial_output.json").exists()
    and (not _viewing_saved_report or _reports_view_filename == "__generating__")
)

# Check if we're on home page
_is_home_page = (
    st.session_state.get("view_mode", "") != "reports"
    and not st.session_state.get("reports_view_filename")
    and not st.session_state.get("reports_view_active")
)

# If on home page, always try to show the latest report (either from current session or saved reports)
if _is_home_page and not _showing_bg_job_table and _status not in ("running", "done"):
    # Check if we already have valid data to display
    _current_key = st.session_state.get("display_output_data_key")
    _has_valid_data = _current_key and _current_key in st.session_state and st.session_state[_current_key] is not None
    
    # If no valid data, load the latest report from saved reports
    if not _has_valid_data:
        _latest_reports = reports_load_index()
        if _latest_reports:
            # Get the most recent report (first in list, sorted by run_at desc)
            _latest_entry = _latest_reports[0]
            _latest_filename = _latest_entry.get("filename", "")
            if _latest_filename and _latest_filename != "__generating__":
                _latest_path = REPORTS_DIR / _latest_filename
                if _latest_path.exists():
                    try:
                        _latest_df = pd.read_excel(_latest_path, engine="openpyxl")
                        _latest_df = _reconstruct_ins_end_marker(_latest_df)
                        _latest_key = f"output_data_latest_{_latest_filename}"
                        st.session_state[_latest_key] = _latest_df
                        st.session_state["display_output_data_key"] = _latest_key
                        st.session_state["display_station_name"] = _latest_entry.get("station", "")
                        # Calculate stats
                        _latest_total_days = 0
                        _latest_total_instructions = _latest_entry.get("total_instructions", 0)
                        if "Date" in _latest_df.columns:
                            _latest_unique_dates = _latest_df["Date"].dropna().astype(str).replace("", pd.NA).dropna().unique()
                            _latest_total_days = len([d for d in _latest_unique_dates if d and d.strip()])
                        if not _latest_total_instructions and "Sum Mus" in _latest_df.columns:
                            _latest_total_instructions = len(_latest_df[_latest_df["Sum Mus"].notna() & (_latest_df["Sum Mus"] != "")])
                        st.session_state["display_stats"] = {
                            "total_days": _latest_total_days,
                            "total_instructions": _latest_total_instructions,
                            "output_rows": _latest_entry.get("row_count", 0) or len(_latest_df),
                        }
                        # Set report title
                        date_f = _latest_entry.get("date_from", "")
                        date_t = _latest_entry.get("date_to", "")
                        if date_f and date_t:
                            st.session_state["report_title"] = f"Back Down Report — {date_f} to {date_t}"
                        elif date_f:
                            st.session_state["report_title"] = f"Back Down Report — {date_f}"
                        else:
                            st.session_state["report_title"] = "Back Down Report"
                    except Exception:
                        pass

# Don't show old report when new generation is running (unless viewing a specific saved report)
_hide_old_report_during_generation = _status == "running" and not _viewing_saved_report

if 'display_output_data_key' in st.session_state and not _showing_bg_job_table and not _hide_old_report_during_generation:
    output_data_key = st.session_state['display_output_data_key']
    station_name_display = st.session_state.get('display_station_name', '')
    
    if output_data_key in st.session_state:
        df_output = st.session_state[output_data_key]
        if df_output is not None and not df_output.empty:
            df_output = df_output.copy()
            df_output = df_output.fillna("").replace("None", "")
            # Make numeric columns Arrow-compatible (float; empty/invalid -> NaN)
            for col in ("DC (MW)", "As per SLDC Scada in MW", "MW as per ramp", "DC , Scada Diff (MW)", "Mus", "Sum Mus", "Diff", "MU", "Sum MU"):
                if col in df_output.columns:
                    df_output[col] = pd.to_numeric(df_output[col], errors="coerce")
            if "DC , Scada Diff (MW)" in df_output.columns:
                df_output["DC , Scada Diff (MW)"] = df_output["DC , Scada Diff (MW)"].apply(
                    lambda x: round(x, 2) if isinstance(x, (int, float)) and pd.notna(x) else x
                )
            if "Sum Mus" in df_output.columns:
                df_output["Sum Mus"] = df_output["Sum Mus"].apply(
                    lambda x: round(x, 3) if isinstance(x, (int, float)) and pd.notna(x) else x
                )
            # Reorder columns to match expected output format (including hidden marker columns)
            expected_cols = ["Date", "From", "To", "DC (MW)", "As per SLDC Scada in MW", "DC , Scada Diff (MW)", "Mus", "Sum Mus", "MW as per ramp", "Diff", "MU", "Sum MU", "_ins_end"]
            df_output = df_output[[c for c in expected_cols if c in df_output.columns]]
        
        processing = st.session_state.get('processing_in_progress', False)
        
        if df_output is not None and not df_output.empty:
            # Show progress caption only when processing
            if processing:
                proc_config = st.session_state.get('processing_config', {})
                total_slots = proc_config.get('total_slots', 1)
                progress_val = min(0.95, 0.6 + 0.3 * len(df_output) / max(1, total_slots))
                st.progress(progress_val)
                current_date = proc_config.get('current_date', '')
                if current_date:
                    st.caption(f"⏳ Processing day {current_date} — {len(df_output)} rows so far")
                else:
                    st.caption(f"⏳ Processing... {len(df_output)} rows so far")
            
            # Stats only when complete
            if not processing and 'display_stats' in st.session_state:
                stats = st.session_state['display_stats']
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Days", stats.get('total_days', 0))
                with col2:
                    st.metric("Total Instructions", stats.get('total_instructions', 0))
                with col3:
                    st.metric("Output Rows", stats.get('output_rows', 0))
                if st.session_state.get("reports_view_active"):
                    url_report_file(st.session_state["reports_view_active"])  # Keep URL: ?view=report&file=...
            
            # Create dynamic table title based on station and date range
            title_parts = ["Calculation sheet for BD and non compliance of", station_name_display]
            
            # Extract date range from report title or use current dates (format: "Back Down Calculator — 01-Jan-2026 to 31-Jan-2026")
            report_title = st.session_state.get('report_title', "Back Down Calculator")
            date_part = ""
            if " — " in report_title:
                date_part = report_title.split(" — ", 1)[1].strip()
            if date_part:
                try:
                    if " to " in date_part:
                        from_date_str, to_date_str = date_part.split(" to ", 1)[0].strip(), date_part.split(" to ", 1)[1].strip()
                        try:
                            from_dt = datetime.strptime(from_date_str, "%d-%b-%Y")
                            to_dt = datetime.strptime(to_date_str, "%d-%b-%Y")
                            if from_dt.year == to_dt.year and from_dt.month == to_dt.month:
                                date_suffix = f"for {from_dt.strftime('%b %y')}"
                            else:
                                date_suffix = f"for {from_dt.strftime('%b %y')} to {to_dt.strftime('%b %y')}"
                        except Exception:
                            date_suffix = f"for {date_part}"
                    else:
                        try:
                            from_dt = datetime.strptime(date_part, "%d-%b-%Y")
                            date_suffix = f"for {from_dt.strftime('%b %y')}"
                        except Exception:
                            date_suffix = f"for {date_part}"
                    title_parts.append(date_suffix)
                except Exception:
                    pass
            
            table_title = " ".join(title_parts)
            
            st.divider()
            st.header(f"📊 {table_title}")
            
            # Prepare day filter options first
            day_idx_key = f"{output_data_key}_day_idx"
            available_dates = []
            if "Date" in df_output.columns:
                available_dates = df_output["Date"].dropna().astype(str).replace("", pd.NA).dropna().unique().tolist()
                available_dates = [d for d in available_dates if d and d.strip()]
                try:
                    available_dates = sorted(available_dates, key=lambda x: datetime.strptime(x, "%d-%b-%Y"))
                except Exception:
                    available_dates = sorted(available_dates)
            
            day_options = ["All Days"] + available_dates if available_dates else ["All Days"]
            current_idx = st.session_state.get(day_idx_key, 0)
            if current_idx >= len(day_options):
                current_idx = 0
            
            # Sync selectbox value with index
            if "day_selectbox" not in st.session_state:
                st.session_state["day_selectbox"] = day_options[current_idx]
            elif st.session_state.get("day_selectbox") != day_options[current_idx]:
                st.session_state["day_selectbox"] = day_options[current_idx]
            
            # Search, Day Filter with Prev/Next, Download - all in same row
            search_key = f"{output_data_key}_search"
            rows_key = f"{output_data_key}_rows_per_page"
            page_key = f"{output_data_key}_page"
            
            col_search, col_prev, col_day, col_next, col_download = st.columns([3, 0.5, 2, 0.5, 1])
            
            with col_search:
                current_search = st.session_state.get(search_key, "")
                search_term = st.text_input(
                    "🔍 Search",
                    value=current_search,
                    placeholder="Search in all columns...",
                    help="Filter rows by searching across all columns",
                    key=search_key
                )
            
            with col_prev:
                st.markdown('<div style="min-height: 1.5rem;">&nbsp;</div>', unsafe_allow_html=True)
                prev_disabled = current_idx <= 1
                def go_prev():
                    new_idx = max(1, st.session_state.get(day_idx_key, 0) - 1)
                    st.session_state[day_idx_key] = new_idx
                    st.session_state["day_selectbox"] = day_options[new_idx]
                st.button("◀", key="day_prev_btn", disabled=prev_disabled, use_container_width=True, on_click=go_prev)
            
            with col_day:
                st.markdown('<div style="min-height: 1.5rem;">&nbsp;</div>', unsafe_allow_html=True)
                def on_day_change():
                    selected = st.session_state["day_selectbox"]
                    st.session_state[day_idx_key] = day_options.index(selected)
                selected_day = st.selectbox(
                    "📅 Filter by Day",
                    options=day_options,
                    key="day_selectbox",
                    on_change=on_day_change,
                    label_visibility="collapsed"
                )
                current_idx = day_options.index(selected_day)
            
            with col_next:
                st.markdown('<div style="min-height: 1.5rem;">&nbsp;</div>', unsafe_allow_html=True)
                next_disabled = current_idx >= len(day_options) - 1
                def go_next():
                    new_idx = min(len(day_options) - 1, st.session_state.get(day_idx_key, 0) + 1)
                    st.session_state[day_idx_key] = new_idx
                    st.session_state["day_selectbox"] = day_options[new_idx]
                st.button("▶", key="day_next_btn", disabled=next_disabled, use_container_width=True, on_click=go_next)
            
            with col_download:
                st.markdown('<div style="min-height: 1.5rem;">&nbsp;</div>', unsafe_allow_html=True)
                # Find the file to download based on various sources
                _dl_file_data = None
                _dl_filename = None
                
                # Priority 1: viewing saved report from Reports page
                viewing_saved = st.session_state.get("reports_view_active")
                if viewing_saved and not _dl_file_data:
                    report_path = REPORTS_DIR / viewing_saved
                    if report_path.exists():
                        try:
                            with open(report_path, "rb") as f:
                                _dl_file_data = f.read()
                            _dl_filename = viewing_saved
                        except Exception:
                            pass
                
                # Priority 2: Extract filename from output_data_key (home page latest/generated)
                if not _dl_file_data and output_data_key:
                    _extracted_filename = None
                    for prefix in ("output_data_latest_", "output_data_home_", "output_data_report_"):
                        if output_data_key.startswith(prefix):
                            _extracted_filename = output_data_key[len(prefix):]
                            break
                    if _extracted_filename:
                        _dl_path = REPORTS_DIR / _extracted_filename
                        if _dl_path.exists():
                            try:
                                with open(_dl_path, "rb") as f:
                                    _dl_file_data = f.read()
                                _dl_filename = _extracted_filename
                            except Exception:
                                pass
                
                # Priority 3: last_output_file_data from session (just generated)
                if not _dl_file_data and 'last_output_file_data' in st.session_state:
                    _dl_file_data = st.session_state['last_output_file_data']
                    _dl_filename = st.session_state.get('last_output_filename', 'output.xlsx')
                
                # Priority 4: last_output_path from session
                if not _dl_file_data and 'last_output_path' in st.session_state:
                    output_path = Path(st.session_state['last_output_path'])
                    if output_path.exists():
                        try:
                            with open(output_path, "rb") as f:
                                _dl_file_data = f.read()
                            _dl_filename = st.session_state.get('last_output_filename', 'output.xlsx')
                        except Exception:
                            pass
                
                # Show download button if we have data
                if _dl_file_data and _dl_filename:
                    st.download_button(
                        label="📥 Download",
                        data=_dl_file_data,
                        file_name=_dl_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_button_output"
                    )
            
            # Apply day filter
            if available_dates and selected_day and selected_day != "All Days":
                # Filter rows where Date matches OR Date is empty (continuation rows belong to previous date)
                # First, forward-fill dates to assign empty rows to their parent date
                df_with_dates = df_output.copy()
                df_with_dates["_date_filled"] = df_with_dates["Date"].replace("", pd.NA).ffill()
                df_day_filtered = df_with_dates[df_with_dates["_date_filled"] == selected_day].drop(columns=["_date_filled"])
            else:
                df_day_filtered = df_output.copy()
            
            # Apply search filter
            if search_term:
                mask = df_day_filtered.astype(str).apply(
                    lambda x: x.str.contains(search_term, case=False, na=False)
                ).any(axis=1)
                df_filtered = df_day_filtered[mask].copy()
            else:
                df_filtered = df_day_filtered.copy()
            
            total_rows = len(df_filtered)
            
            if AGGrid_AVAILABLE:
                # Page size options: 20, 50, 100, 500, and All (show all = total_rows); default All; display "ALL" in selector
                gb = GridOptionsBuilder.from_dataframe(df_filtered)
                page_size_options = sorted(set([20, 50, 100, 500, total_rows])) if total_rows > 0 else [20]
                default_page_size = total_rows if total_rows > 0 else 20
                gb.configure_pagination(
                    paginationAutoPageSize=False,
                    paginationPageSize=default_page_size,
                )
                gb.configure_grid_options(
                    paginationPageSizeSelector=page_size_options,
                    onFirstDataRendered=JsCode(
                        f"""
                        function(params) {{
                            var allVals = ['{PAGE_SIZE_ALL}', '{total_rows}'];
                            function replacePageSizeText(root) {{
                                try {{
                                    var walk = document.createTreeWalker(root, NodeFilter.SHOW_TEXT);
                                    var n;
                                    while ((n = walk.nextNode())) {{
                                        var t = n.textContent.trim();
                                        if (allVals.indexOf(t) !== -1) n.textContent = 'ALL';
                                    }}
                                    if (root.querySelectorAll) root.querySelectorAll('select option').forEach(function(opt) {{
                                        if (allVals.indexOf(opt.value) !== -1) opt.textContent = 'ALL';
                                    }});
                                }} catch (e) {{}}
                            }}
                            function run() {{
                                var el = params.api.getGridElement();
                                if (el) {{
                                    var root = el.closest('.ag-root-wrapper') || el.closest('.ag-root') || el;
                                    if (root) replacePageSizeText(root);
                                }}
                                replacePageSizeText(document.body);
                            }}
                            setTimeout(run, 100);
                            setTimeout(run, 500);
                        }}
                        """
                    ),
                )
                gb.configure_side_bar()
                gb.configure_default_column(
                    sortable=True,
                    filterable=True,
                    resizable=True,
                    editable=False
                )
                gb.configure_selection('single')
                
                # Cell styling: Yellow highlight for Date (when not empty) and To (end of instruction)
                # Date column: highlight yellow when cell has a value (first row of each date)
                date_cell_style = JsCode("""
                function(params) {
                    if (params.value && params.value.toString().trim() !== '') {
                        return {'backgroundColor': '#FFFF00', 'fontWeight': 'bold'};
                    }
                    return null;
                }
                """)
                gb.configure_column("Date", cellStyle=date_cell_style)
                
                # To column: highlight yellow when it's the end of an instruction (from input file)
                # Uses the hidden _ins_end marker set during data generation
                to_cell_style = JsCode("""
                function(params) {
                    var rowData = params.data;
                    var insEnd = rowData['_ins_end'];
                    // Highlight if this row is marked as instruction end
                    // Handle various data type representations of boolean true
                    if (insEnd === true || insEnd === 1 || insEnd === 'True' || insEnd === 'true' || insEnd === 'TRUE') {
                        return {'backgroundColor': '#FFFF00', 'fontWeight': 'bold'};
                    }
                    return null;
                }
                """)
                gb.configure_column("To", cellStyle=to_cell_style)
                
                # Hide the _ins_end marker column
                gb.configure_column("_ins_end", hide=True)
                
                gridOptions = gb.build()
                # Height based on row count, capped at max
                _display_rows = min(total_rows, 50) if total_rows > 0 else 20
                AgGrid(
                    df_filtered,
                    gridOptions=gridOptions,
                    height=table_height(_display_rows),
                    width='100%',
                    theme='streamlit',
                    update_mode=GridUpdateMode.NO_UPDATE,
                    data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                    allow_unsafe_jscode=True
                )
            else:
                # Fallback: manual pagination (no "Rows per page" selectbox)
                rows_per_page = 20
                total_pages = (total_rows + rows_per_page - 1) // rows_per_page if total_rows > 0 else 1
                current_page = st.session_state.get(page_key, 1)
                if current_page > total_pages:
                    current_page = 1
                page_num = current_page
                start_idx = (page_num - 1) * rows_per_page
                end_idx = start_idx + rows_per_page
                df_display = df_filtered.iloc[start_idx:end_idx].copy()
                
                st.dataframe(
                    df_display,
                    width='stretch',
                    height=table_height(len(df_display)),
                    hide_index=True
                )
                if total_pages > 1:
                    st.caption(f"Showing rows {start_idx + 1} to {min(end_idx, total_rows)} of {total_rows}")
                else:
                    st.caption(f"Showing all {total_rows} rows")

# Generate button - triggered from sidebar
_viewing_report = bool(st.session_state.get("reports_view_active"))
run_generate = False
if not _viewing_report and _status != "running":
    # Check if sidebar generate button was clicked
    run_generate = st.session_state.pop('_sidebar_generate_clicked', False) or st.session_state.pop('_run_continue_processing', False)
if run_generate:
    # Start report generation in a background thread so it continues when user navigates away
    try:
        temp_base = Path(tempfile.gettempdir()) / "electrical_app"
        temp_base.mkdir(parents=True, exist_ok=True)
        run_id = str(uuid.uuid4())[:8]
        temp_path = temp_base / run_id
        temp_path.mkdir(exist_ok=True)

        instructions_path = temp_path / instructions_file.name
        with open(instructions_path, "wb") as f:
            f.write(instructions_file.getbuffer())
        dc_name = ""
        if dc_file:
            dc_path = temp_path / dc_file.name
            with open(dc_path, "wb") as f:
                f.write(dc_file.getbuffer())
            dc_name = dc_file.name

        job_data = {
            "status": "running",
            "temp_path": str(temp_path),
            "instructions_name": instructions_file.name,
            "dc_name": dc_name,
            "bd_folder_path": bd_folder_path or "",
            "sheet_name": sheet_name or "",
            "column_name": column_name or "Name of the station",
            "station_name": station_name or "",
            "header_rows": header_rows,
            "data_only": data_only,
            "bd_sheet": bd_sheet or "",
            "scada_column": scada_column or "",
            "report_title": st.session_state.get("report_title", "Back Down Calculator"),
            "ramp_up_5": _parse_float(st.session_state.get("ramp_up_5_input", "15"), 15),
            "ramp_up_10": _parse_float(st.session_state.get("ramp_up_10_input", "27.5"), 27.5),
            "ramp_up_15": _parse_float(st.session_state.get("ramp_up_15_input", "40"), 40),
            "ramp_down_5": _parse_float(st.session_state.get("ramp_down_5_input", "15"), 15),
            "ramp_down_10": _parse_float(st.session_state.get("ramp_down_10_input", "27.5"), 27.5),
            "ramp_down_15": _parse_float(st.session_state.get("ramp_down_15_input", "40"), 40),
            "created_at": datetime.now().isoformat(),
            "progress_pct": 0,
            "processed_slots": 0,
            "total_slots": 0,
        }
        background_write_job(job_data)
        thread = threading.Thread(target=_run_report_generation_worker, args=(job_data,), daemon=True)
        thread.start()
        st.success("Report generation started in the background. You can switch to Reports or other pages.")
        st.rerun()
    except Exception as e:
        st.error(f"Failed to start report generation: {str(e)}")
        if verbose:
            import traceback
            st.code(traceback.format_exc())

