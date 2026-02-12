#!/usr/bin/env python3
r"""
Read an XLSX file: search for sheet by name, refer to column by name, optionally find rows by column value.

Usage:
  python xlsx_sheet_column.py <xlsx_path> <sheet_name> <column_name> [column_value] [options]

Example:
  python xlsx_sheet_column.py "input/BD  LR  01-01-2026.xlsx" "SCADA Grid" "AP_UI.MW"
  python xlsx_sheet_column.py "input/BD  LR  01-01-2026.xlsx" "SCADA Grid" "Time" "2026-01-01 00:15"
  python xlsx_sheet_column.py "input/jan 2026.xlsx" HNPCL "Date" "2026-01-01" --data-only
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def find_sheet_by_name(wb, sheet_name):
    """
    Search for a sheet in the workbook by name (case-insensitive, partial match).
    Returns (sheet, exact_name) or (None, None).
    """
    target = sheet_name.strip().lower()
    if not target:
        return None, None
    # Exact match first
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name], name
    # Partial match (target contained in sheet name)
    for name in wb.sheetnames:
        if target in name.strip().lower():
            return wb[name], name
    return None, None


def find_column_by_name(ws, column_name, max_header_rows=10):
    """
    Search for a column in the sheet by header name (case-insensitive, partial match).
    Scans first max_header_rows for header row.
    Returns (1-based column index, header_row) or (None, None).
    """
    target = column_name.strip().lower()
    if not target:
        return None, None
    for row_num in range(1, min(max_header_rows + 1, ws.max_row + 1)):
        for col_idx, cell in enumerate(ws[row_num], start=1):
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            if val == target:
                return col_idx, row_num
            if target in val:
                return col_idx, row_num
    return None, None


def main():
    parser = argparse.ArgumentParser(
        description="Read XLSX file, find sheet by name, refer to column by name."
    )
    parser.add_argument(
        "xlsx_path",
        type=Path,
        help="Path to the XLSX file",
    )
    parser.add_argument(
        "sheet_name",
        help="Sheet name to search for in the file (partial match supported)",
    )
    parser.add_argument(
        "column_name",
        help="Column name to search for in the sheet (partial match supported)",
    )
    parser.add_argument(
        "column_value",
        nargs="?",
        default=None,
        help="Column value to match: print the full row(s) where this column equals this value",
    )
    parser.add_argument(
        "--list",
        action="store_true",
        help="List values in the column (from first data row to end)",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=20,
        help="Max rows to show when --list (default: 20)",
    )
    parser.add_argument(
        "--header-rows",
        type=int,
        default=10,
        help="Max header rows to scan for column name (default: 10)",
    )
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="Read with data_only=True (evaluated values, not formulas)",
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx_path
    if not xlsx_path.is_file():
        print(f"Error: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=args.data_only)

    # Find sheet
    sheet, matched_name = find_sheet_by_name(wb, args.sheet_name)
    if sheet is None:
        print(f"Error: No sheet matching '{args.sheet_name}' in {xlsx_path.name}", file=sys.stderr)
        print(f"Available sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
        wb.close()
        sys.exit(1)

    print(f"Sheet: {matched_name}")

    # Find column
    col_idx, header_row = find_column_by_name(sheet, args.column_name, max_header_rows=args.header_rows)
    if col_idx is None:
        print(f"Error: No column matching '{args.column_name}' in sheet '{matched_name}'", file=sys.stderr)
        # Show first row as hint
        first_row = [str(cell.value)[:30] if cell.value else "" for cell in sheet[header_row or 1]]
        print(f"First row sample: {first_row[:15]}", file=sys.stderr)
        wb.close()
        sys.exit(1)

    col_letter = openpyxl.utils.get_column_letter(col_idx)
    print(f"Column: '{args.column_name}' -> column {col_letter} (index {col_idx}), header row {header_row}")

    data_start = (header_row or 1) + 1

    if args.column_value is not None:
        # Find and print rows where column value matches
        target_val = args.column_value.strip()
        matched_rows = []
        for row_num in range(data_start, sheet.max_row + 1):
            cell = sheet.cell(row=row_num, column=col_idx)
            val = cell.value
            if val is None:
                if target_val == "" or target_val.lower() == "none":
                    matched_rows.append(row_num)
                continue
            val_str = str(val).strip()
            if val_str == target_val:
                matched_rows.append(row_num)
            elif target_val.lower() in val_str.lower():
                matched_rows.append(row_num)
        if not matched_rows:
            print(f"\nNo row where column '{args.column_name}' equals '{args.column_value}'")
        else:
            print(f"\nMatched {len(matched_rows)} row(s) where column = '{args.column_value}':")
            for row_num in matched_rows:
                row_cells = [sheet.cell(row=row_num, column=c).value for c in range(1, sheet.max_column + 1)]
                row_preview = [str(v)[:40] if v is not None else "" for v in row_cells[:15]]
                print(f"  Row {row_num}: {row_preview}")
                if sheet.max_column > 15:
                    print(f"           ... ({sheet.max_column} columns total)")

    elif args.list:
        print(f"\nValues (rows {data_start} to min(end, {data_start + args.max_rows - 1})):")
        for row_num in range(data_start, min(data_start + args.max_rows, sheet.max_row + 1)):
            cell = sheet.cell(row=row_num, column=col_idx)
            val = cell.value
            print(f"  Row {row_num}: {val}")

    wb.close()


if __name__ == "__main__":
    main()
