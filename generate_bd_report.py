#!/usr/bin/env python3
"""
Automated generation of Back Down and Non-compliance report.

Reads input files:
1. Reference calculation sheet - Contains instruction periods (extracted from Date column) and format template
2. HNPCL revised DC file - DC (MW) values (INPUT)
3. Daily BD LR files - SCADA readings (INPUT)

Generates output Excel file matching reference format with fresh data from input files.
"""

import argparse
import datetime
import sys
from pathlib import Path
from datetime import datetime as dt, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_time(value):
    """Parse time from various formats."""
    if value is None:
        return None
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, str):
        # Try parsing HH:MM or HH:MM:SS
        for fmt in ["%H:%M", "%H:%M:%S", "%H.%M"]:
            try:
                return dt.strptime(value, fmt).time()
            except ValueError:
                continue
    return None


def parse_date(value):
    """Parse date from various formats."""
    if value is None:
        return None
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, str):
        # Try common date formats
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"]:
            try:
                return dt.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def time_to_minutes(t):
    """Convert time to minutes since midnight."""
    if t is None:
        return None
    if isinstance(t, datetime.datetime):
        t = t.time()
    if isinstance(t, datetime.time):
        return t.hour * 60 + t.minute
    return None


def minutes_to_time(minutes):
    """Convert minutes since midnight to time."""
    if minutes is None:
        return None
    h = minutes // 60
    m = minutes % 60
    return datetime.time(h, m)


def floor_to_15(minutes):
    """Floor minutes to previous 15-minute slot."""
    if minutes is None:
        return None
    return (minutes // 15) * 15


def calculate_variable_a(from_time, to_time):
    """
    Calculate variable A based on instruction start time.
    From handwritten notes:
    - 15 min duration: A = 40
    - 10 min duration: A = 27.5
    - 5 min duration: A = 15
    """
    from_min = time_to_minutes(from_time)
    to_min = time_to_minutes(to_time)
    
    if from_min is None or to_min is None:
        return 40  # Default
    
    # Handle overnight
    if to_min < from_min:
        to_min += 24 * 60
    
    duration_min = to_min - from_min
    
    if duration_min >= 15:
        return 40
    elif duration_min >= 10:
        return 27.5
    elif duration_min >= 5:
        return 15
    else:
        # Pro-rated for less than 5 minutes
        return (duration_min / 15) * 40


def generate_15min_slots(from_time, to_time):
    """Generate 15-minute slots between from_time and to_time."""
    start_min = time_to_minutes(from_time)
    end_min = time_to_minutes(to_time)
    
    if start_min is None or end_min is None:
        return []
    
    start_slot = floor_to_15(start_min)
    end_slot = floor_to_15(end_min)
    
    if start_slot == end_slot:
        return [(minutes_to_time(start_slot), minutes_to_time((start_slot + 15) % (24 * 60)))]
    
    # Handle overnight
    if start_slot > end_slot:
        end_slot += 24 * 60
    
    slots = []
    m = start_slot
    while m <= end_slot:
        from_m = m % (24 * 60)
        to_m = (m + 15) % (24 * 60)
        slots.append((minutes_to_time(from_m), minutes_to_time(to_m)))
        m += 15
    
    return slots


def extract_instruction_periods_from_calc_sheet(calc_file):
    """
    Extract instruction periods from calculation sheet.
    Instruction periods are marked by dates in Column B - each date marks start of a new period.
    """
    wb = openpyxl.load_workbook(calc_file, read_only=True, data_only=True)
    ws = wb["New method"]
    
    periods = []
    data_start_row = 5
    
    current_date = None
    period_start_row = None
    period_start_from = None
    
    for row_num in range(data_start_row, ws.max_row + 1):
        date_cell = ws[f'B{row_num}']
        from_cell = ws[f'C{row_num}']
        to_cell = ws[f'D{row_num}']
        
        # Check if this is start of new period (has date)
        if date_cell.value:
            # Save previous period if exists
            if period_start_row is not None and period_start_from is not None:
                # Get last "To" time from previous row
                prev_to_cell = ws[f'D{row_num - 1}']
                if prev_to_cell.value:
                    prev_to = parse_time(prev_to_cell.value)
                    period_from = parse_time(period_start_from)
                    if prev_to and period_from and current_date:
                        periods.append((current_date, period_from, prev_to))
            
            # Start new period
            current_date = parse_date(date_cell.value)
            period_start_from = from_cell.value  # Keep as raw value, parse later
            period_start_row = row_num
        
        # Check if we've reached end of sheet
        if row_num == ws.max_row:
            # Save last period
            if period_start_row is not None and period_start_from is not None:
                last_to_cell = ws[f'D{row_num}']
                if last_to_cell.value:
                    last_to = parse_time(last_to_cell.value)
                    period_from = parse_time(period_start_from)
                    if last_to and period_from and current_date:
                        periods.append((current_date, period_from, last_to))
    
    wb.close()
    return periods


def get_dc_value(dc_file, date, time_slot):
    """Get DC (MW) value from HNPCL Revised DC file for given date and time."""
    # Format date for sheet name (e.g., "01.01.2026")
    base_sheet_name = date.strftime("%d.%m.%Y")
    
    wb = openpyxl.load_workbook(dc_file, read_only=True)
    
    # Try to find matching sheet (handle variations with/without spaces)
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip() == base_sheet_name or name.strip().startswith(base_sheet_name):
            sheet_name = name
            break
    
    if sheet_name is None:
        wb.close()
        return None
    
    ws = wb[sheet_name]
    
    # Find matching time block
    target_min = time_to_minutes(time_slot)
    
    for row_num in range(3, ws.max_row + 1):
        from_cell = ws[f'B{row_num}']
        dc_cell = ws[f'E{row_num}']  # Final Revision column
        
        if from_cell.value is None or dc_cell.value is None:
            continue
        
        from_time = parse_time(from_cell.value)
        if from_time:
            from_min = time_to_minutes(from_time)
            # Match within same 15-minute slot
            if from_min is not None and (from_min // 15) == (target_min // 15):
                dc_value = dc_cell.value
                wb.close()
                return float(dc_value) if isinstance(dc_value, (int, float)) else None
    
    wb.close()
    return None


def get_scada_value(bd_file, date, time_slot):
    """Get SCADA value from Daily BD LR file for given date and time."""
    wb = openpyxl.load_workbook(bd_file, read_only=True)
    
    if "SCADA Grid" not in wb.sheetnames:
        wb.close()
        return None
    
    ws = wb["SCADA Grid"]
    
    # Create datetime for matching
    target_dt = dt.combine(date, time_slot)
    
    # SCADA Grid has time in Column A (row 4 onwards)
    for row_num in range(4, ws.max_row + 1):
        time_cell = ws[f'A{row_num}']
        scada_cell = ws[f'D{row_num}']  # SCHED_PG.SYSTEM.AP_UI.MW column
        
        if time_cell.value is None:
            continue
        
        if isinstance(time_cell.value, datetime.datetime):
            if time_cell.value.replace(second=0, microsecond=0) == target_dt.replace(second=0, microsecond=0):
                scada_value = scada_cell.value
                wb.close()
                return float(scada_value) if isinstance(scada_value, (int, float)) else None
    
    wb.close()
    return None


def find_daily_bd_file(input_dir, date):
    """Find the Daily BD LR file for given date."""
    # Format: "BD  LR  DD-MM-YYYY.xlsx" or "BD LR_MBED DD-MM-YYYY.xlsx"
    date_str = date.strftime("%d-%m-%Y")
    
    patterns = [
        f"BD  LR  {date_str}.xlsx",
        f"BD LR_MBED {date_str}.xlsx",
        f"BD LR {date_str}.xlsx",
    ]
    
    for pattern in patterns:
        file_path = input_dir / pattern
        if file_path.exists():
            return file_path
    
    return None


def generate_output(input_dir, reference_file, dc_file, output_file):
    """
    Generate the output Excel file.
    
    Uses reference_file (calculation sheet) as template to:
    1. Extract instruction periods (Date, From, To)
    2. Understand output format and structure
    3. Generate new output with fresh data from input files
    """
    print(f"Using reference file: {reference_file.name}")
    print(f"Extracting instruction periods from reference file...")
    periods = extract_instruction_periods_from_calc_sheet(reference_file)
    print(f"Found {len(periods)} instruction periods")
    
    # Generate all 15-minute slots for all periods
    all_rows = []
    
    for date, from_time, to_time in periods:
        slots = generate_15min_slots(from_time, to_time)
        variable_a = calculate_variable_a(from_time, to_time)
        
        for idx, (slot_from, slot_to) in enumerate(slots):
            is_first = (idx == 0)
            all_rows.append({
                'date': date if is_first else None,
                'from_time': slot_from,
                'to_time': slot_to,
                'variable_a': variable_a if is_first else None,
                'is_first_in_period': is_first,
                'period_start': from_time,  # Store original period start for reference
                'period_end': to_time
            })
    
    print(f"Generated {len(all_rows)} data rows")
    
    # Fetch DC and SCADA values
    print("Fetching DC and SCADA values...")
    current_date = None
    for row in all_rows:
        # Propagate date forward
        if row['date']:
            current_date = row['date']
        elif current_date:
            row['date'] = current_date
        else:
            print(f"Warning: Row without date: {row}")
            continue
        
        # Get DC value
        dc_value = get_dc_value(dc_file, current_date, row['from_time'])
        row['dc_mw'] = dc_value
        
        # Get SCADA value
        bd_file = find_daily_bd_file(input_dir, current_date)
        if bd_file:
            scada_value = get_scada_value(bd_file, current_date, row['from_time'])
            row['scada_mw'] = scada_value
        else:
            row['scada_mw'] = None
    
    # Variable A is already calculated and stored in all_rows
    
    # Create output workbook
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Back down and Non compliance"
    
    # Define styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Extract title from reference file
    ref_wb = openpyxl.load_workbook(reference_file, read_only=True)
    ref_ws = ref_wb["New method"]
    ref_title = ref_ws['B2'].value if ref_ws['B2'].value else "Back down and Non compliance of HNPCL"
    ref_wb.close()
    
    # Row 1: Empty (matches reference)
    # Row 2: Main title (matches reference: B2:J2)
    output_sheet.merge_cells('B2:J2')
    title_cell = output_sheet['B2']
    title_cell.value = ref_title
    title_cell.font = title_font
    title_cell.alignment = center_align
    
    # Row 3: Category headers (matches reference: B3:I3 and J3:M3)
    output_sheet.merge_cells('B3:I3')
    backdown_header = output_sheet['B3']
    backdown_header.value = "Back down"
    backdown_header.font = header_font
    backdown_header.alignment = center_align
    backdown_header.border = thin_border
    
    output_sheet.merge_cells('J3:M3')
    noncompliance_header = output_sheet['J3']
    noncompliance_header.value = "Non compliance"
    noncompliance_header.font = header_font
    noncompliance_header.alignment = center_align
    noncompliance_header.border = thin_border
    
    # Row 4: Column headers (matches reference - starts from Column B)
    headers_backdown = ["Date", "From", "To", "DC\n(MW)", "As per SLDC Scada in MW", "Diff (MW)", "Mus"]
    headers_noncompliance = ["MW as per ramp", "Diff ", "MU"]
    
    for col_idx, header in enumerate(headers_backdown, start=2):  # Start from Column B (index 2)
        cell = output_sheet.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        if header == "As per SLDC Scada in MW":
            cell.fill = yellow_fill
    
    for col_idx, header in enumerate(headers_noncompliance, start=10):  # Start from Column J (index 10)
        cell = output_sheet.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data rows starting from row 5 (matches reference - data starts at row 5)
    for row_idx, row_data in enumerate(all_rows, start=5):
        # Column B: Date
        if row_data['date']:
            output_sheet.cell(row=row_idx, column=2).value = row_data['date']
        
        # Column C: From
        if row_data['from_time']:
            output_sheet.cell(row=row_idx, column=3).value = row_data['from_time'].strftime("%H:%M")
        
        # Column D: To
        if row_data['to_time']:
            output_sheet.cell(row=row_idx, column=4).value = row_data['to_time'].strftime("%H:%M")
        
        # Column E: DC (MW)
        if row_data['dc_mw'] is not None:
            output_sheet.cell(row=row_idx, column=5).value = row_data['dc_mw']
        
        # Column F: As per SLDC Scada in MW
        if row_data['scada_mw'] is not None:
            output_sheet.cell(row=row_idx, column=6).value = row_data['scada_mw']
        
        # Column G: Diff (MW) = E - F
        if row_data['dc_mw'] is not None and row_data['scada_mw'] is not None:
            output_sheet.cell(row=row_idx, column=7).formula = f"=E{row_idx}-F{row_idx}"
        
        # Column H: Mus = G / 4000
        output_sheet.cell(row=row_idx, column=8).formula = f"=G{row_idx}/4000"
        
        # Column J: MW as per ramp (with formula)
        if row_data['is_first_in_period'] and row_data['variable_a'] is not None:
            # First cell: =E{row} - A
            variable_a = row_data['variable_a']
            output_sheet.cell(row=row_idx, column=10).formula = f"=E{row_idx}-{variable_a}"
        elif not row_data['is_first_in_period']:
            # Subsequent: =MAX(J{prev}-40, 270)
            prev_row = row_idx - 1
            output_sheet.cell(row=row_idx, column=10).formula = f"=MAX(J{prev_row}-40,270)"
        
        # Column K: Diff = F - J
        output_sheet.cell(row=row_idx, column=11).formula = f"=F{row_idx}-J{row_idx}"
        
        # Column L: MU = IF(K/4000>0, K/4000, 0)
        output_sheet.cell(row=row_idx, column=12).formula = f"=IF(K{row_idx}/4000>0,K{row_idx}/4000,0)"
        
        # Apply borders
        for col in range(2, 13):
            output_sheet.cell(row=row_idx, column=col).border = thin_border
    
    # Adjust column widths
    output_sheet.column_dimensions['B'].width = 12  # Date
    output_sheet.column_dimensions['C'].width = 8   # From
    output_sheet.column_dimensions['D'].width = 8   # To
    output_sheet.column_dimensions['E'].width = 10  # DC (MW)
    output_sheet.column_dimensions['F'].width = 25  # As per SLDC Scada in MW
    output_sheet.column_dimensions['G'].width = 12  # Diff (MW)
    output_sheet.column_dimensions['H'].width = 10  # Mus
    output_sheet.column_dimensions['J'].width = 15  # MW as per ramp
    output_sheet.column_dimensions['K'].width = 10  # Diff
    output_sheet.column_dimensions['L'].width = 10  # MU
    
    # Set row heights (matches reference structure)
    output_sheet.row_dimensions[2].height = 25  # Title row
    output_sheet.row_dimensions[3].height = 20  # Category headers
    output_sheet.row_dimensions[4].height = 30  # Column headers
    
    # Save
    output_wb.save(output_file)
    print(f"\nOutput saved to: {output_file}")
    print(f"Total rows created: {len(all_rows)}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Back Down and Non-compliance report from input files"
    )
    parser.add_argument(
        "input_directory",
        help="Directory containing input files (calculation sheet, HNPCL revised DC file, Daily BD LR files)"
    )
    parser.add_argument(
        "--reference",
        help="Path to reference calculation sheet file (default: searches in input_directory)",
        default=None
    )
    parser.add_argument(
        "--dc-file",
        help="Path to HNPCL revised DC file (default: searches in input_directory)",
        default=None
    )
    parser.add_argument(
        "--output",
        help="Output file path (default: input_directory/Back_down_and_Non_compliance_output.xlsx)",
        default=None
    )
    
    args = parser.parse_args()
    
    input_dir = Path(args.input_directory)
    if not input_dir.is_dir():
        print(f"Error: Directory does not exist: {input_dir}", file=sys.stderr)
        sys.exit(1)
    
    # Find reference file (calculation sheet - contains instruction periods and format)
    if args.reference:
        reference_file = Path(args.reference)
    else:
        reference_files = list(input_dir.glob("*calculation*sheet*.xlsx"))
        if not reference_files:
            print(f"Error: No reference calculation sheet found in {input_dir}", file=sys.stderr)
            print(f"Looking for files matching: *calculation*sheet*.xlsx", file=sys.stderr)
            sys.exit(1)
        reference_file = reference_files[0]
    
    # Find DC file
    if args.dc_file:
        dc_file = Path(args.dc_file)
    else:
        dc_files = list(input_dir.glob("*HNPCL*revised*DC*.xlsx"))
        if not dc_files:
            print(f"Error: No HNPCL revised DC file found in {input_dir}", file=sys.stderr)
            sys.exit(1)
        dc_file = dc_files[0]
    
    # Output file
    if args.output:
        output_file = Path(args.output)
    else:
        output_file = input_dir / "Back_down_and_Non_compliance_output.xlsx"
    
    print(f"Input directory: {input_dir}")
    print(f"Reference file: {reference_file}")
    print(f"DC file: {dc_file}")
    print(f"Output file: {output_file}")
    print()
    
    generate_output(input_dir, reference_file, dc_file, output_file)


if __name__ == "__main__":
    main()
